#!/usr/bin/env python3
"""
Import historical courier worksheet files into Stallion Courier manifests.

Usage:
  python scripts/import_courier_history.py \
    --source /path/to/worksheets \
    --arrival-date 2026-04-15 \
    --exch-rate 6.78

Notes:
- Expects Worksheet v3-like layout where line items start at row 8:
  A line_no, B hawb, C shipper, D importer, E description,
  F packages, G weight_kg, H thn, I duty rate, J cost_usd, K freight_usd
- Stops reading when description+thn+cost are all empty for a row.
- Skips duplicates by manifest_no.
"""
from __future__ import annotations

import argparse
import re
import sys
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
BACKEND = ROOT / "backend"
sys.path.insert(0, str(BACKEND))

from app.services import courier_service  # type: ignore


def _to_num(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except Exception:
        return default


def _manifest_no_from_filename(p: Path) -> str:
    m = re.search(r"(106-\d{6,})", p.name)
    if m:
        return m.group(1)
    return p.stem


def _iter_lines(ws):
    row = 8
    while row < 5000:
        line_no = ws[f"A{row}"].value
        hawb = ws[f"B{row}"].value
        shipper = ws[f"C{row}"].value
        importer = ws[f"D{row}"].value
        description = ws[f"E{row}"].value
        packages = ws[f"F{row}"].value
        weight = ws[f"G{row}"].value
        thn = ws[f"H{row}"].value
        cost = ws[f"J{row}"].value
        freight = ws[f"K{row}"].value

        if (description in (None, "")) and (thn in (None, "")) and (cost in (None, "")):
            break

        yield {
            "line_no": int(_to_num(line_no, default=row - 7)),
            "hawb": "" if hawb is None else str(hawb).strip(),
            "shipper": "" if shipper is None else str(shipper).strip(),
            "importer": "" if importer is None else str(importer).strip(),
            "description": "" if description is None else str(description).strip(),
            "packages": int(_to_num(packages, default=1)),
            "weight_kg": _to_num(weight, default=0),
            "thn": "" if thn is None else str(thn).replace(".", "").strip(),
            "cost_usd": _to_num(cost, default=0),
            "freight_usd": _to_num(freight, default=0),
        }
        row += 1


def import_file(path: Path, arrival_date: str, exch_rate: float, dry_run: bool = False) -> tuple[bool, str]:
    manifest_no = _manifest_no_from_filename(path)

    existing = courier_service.list_manifests()
    if any(m.get("manifest_no") == manifest_no for m in existing):
        return False, f"skip duplicate manifest_no={manifest_no} ({path.name})"

    wb = load_workbook(path, data_only=False)
    ws = wb.active

    lines = list(_iter_lines(ws))
    if not lines:
        return False, f"skip no lines parsed ({path.name})"

    if dry_run:
        return True, f"dry-run import {manifest_no}: {len(lines)} lines"

    m = courier_service.create_manifest(
        {
            "manifest_no": manifest_no,
            "arrival_date": arrival_date,
            "exch_rate": exch_rate,
            "cargo_reporter": "TTPOST",
            "notes": f"Imported from {path.name}",
        }
    )

    for ln in lines:
        payload = {
            "hawb": ln["hawb"],
            "shipper": ln["shipper"],
            "importer": ln["importer"],
            "description": ln["description"],
            "packages": ln["packages"],
            "weight_kg": ln["weight_kg"],
            "thn": ln["thn"],
            "cost_usd": ln["cost_usd"],
            "freight_usd": ln["freight_usd"],
        }
        if not payload["thn"]:
            payload["auto_classify"] = True
            courier_service.add_line_with_auto_thn(m["id"], payload)
        else:
            courier_service.add_line(m["id"], payload)

    return True, f"imported {manifest_no}: {len(lines)} lines"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--source", required=True, help="Folder containing worksheet xlsx files")
    ap.add_argument("--pattern", default="*v3*.xlsx", help="Glob pattern")
    ap.add_argument("--arrival-date", required=True, help="Arrival date YYYY-MM-DD")
    ap.add_argument("--exch-rate", required=True, type=float, help="Exchange rate")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    src = Path(args.source)
    files = sorted(src.glob(args.pattern))
    if not files:
        print("No files found")
        return 1

    ok = 0
    for f in files:
        try:
            success, msg = import_file(f, args.arrival_date, args.exch_rate, dry_run=args.dry_run)
            print(msg)
            if success:
                ok += 1
        except Exception as e:
            print(f"error {f.name}: {e}")

    print(f"Done. Successful imports: {ok}/{len(files)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
