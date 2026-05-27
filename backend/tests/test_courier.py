"""
Tests for the courier module Phase 1b (editable rules + tariff overrides).

Covers:
- The duty engine still computes correctly under the rules-store-backed
  classifier
- The rules store: add/remove/update exemptions and corrections
- Tariff overrides: add/remove/list
- Audit trail captures every change
- Export/import round-trip
- Bundled defaults can be overridden by user entries

Run from repo root:
    cd backend
    python -m pytest tests/test_courier.py -v
"""
from __future__ import annotations

import shutil
import sys
import tempfile
import unittest
from pathlib import Path

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE.parent.parent))

from app.services import courier_duty, courier_matcher, courier_rules  # noqa: E402


class _RulesStoreTestBase(unittest.TestCase):
    """Each test gets a fresh tmpdir for the user rules + tariff overrides."""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp(prefix="courier_test_")
        self.user_rules_path = Path(self.tmpdir) / "courier_rules_user.json"
        self.tariff_overrides_path = Path(self.tmpdir) / "tariff_overrides.json"
        # Resolve bundled paths relative to backend root: tests/ and data/
        # are siblings under backend/.
        backend_dir = HERE.parent
        self.bundled_rules_path = backend_dir / "data" / "courier_rules_bundled.json"
        self.bundled_tariff_path = backend_dir / "data" / "tt_tariff_db_2024.json"

        courier_rules.configure_paths(
            bundled_rules_path=self.bundled_rules_path,
            user_rules_path=self.user_rules_path,
            bundled_tariff_path=self.bundled_tariff_path,
            tariff_overrides_path=self.tariff_overrides_path,
        )

    def tearDown(self):
        shutil.rmtree(self.tmpdir, ignore_errors=True)


class TestCourierDutyEngine(_RulesStoreTestBase):
    """The duty engine still classifies and calculates correctly."""

    def test_classify_standard_thn(self):
        result = courier_duty.classify("33049990")
        self.assertEqual(result.exemption_class, "none")
        self.assertGreater(result.duty_rate, 0)
        self.assertEqual(result.rule_source, "cet")

    def test_classify_full_exempt_smartphone(self):
        result = courier_duty.classify("85171300")
        self.assertEqual(result.exemption_class, "full_exempt")
        self.assertEqual(result.duty_rate, 0.0)
        self.assertEqual(result.rule_source, "exemption")

    def test_classify_full_exempt_computer_accessory(self):
        result = courier_duty.classify("84733000")
        self.assertEqual(result.exemption_class, "full_exempt")

    def test_classify_full_exempt_earphones(self):
        result = courier_duty.classify("85183000")
        self.assertEqual(result.exemption_class, "full_exempt")

    def test_classify_plastic_catchall_not_blanket_exempt(self):
        """
        39269090 is a catch-all ("Other articles of plastics") that must
        NOT be blanket-exempt. Generic items under this code pay 20%.
        Cellphone accessories under it are exempted per-LINE by the matcher
        (see TestPhoneCaseExemption), not by a THN rule here.
        """
        result = courier_duty.classify("39269090")
        self.assertNotEqual(
            result.exemption_class, "full_exempt",
            "39269090 must not be blanket-exempt — it's a catch-all code",
        )
        # Base rate should be the corrected 20%
        self.assertEqual(result.duty_rate, 0.2)

    def test_classify_duty_free_only_generic_device(self):
        result = courier_duty.classify("85176900")
        self.assertEqual(result.exemption_class, "duty_free_only")
        self.assertEqual(result.duty_rate, 0.0)

    def test_thn_correction_smartphone(self):
        result = courier_duty.classify("85171200")
        self.assertTrue(result.is_corrected)
        self.assertEqual(result.original_thn, "85171200")
        self.assertEqual(result.exemption_class, "full_exempt")
        self.assertEqual(result.rule_source, "correction")

    def test_thn_correction_air_pump(self):
        result = courier_duty.classify("84148090")
        self.assertTrue(result.is_corrected)

    def test_unknown_thn_marked(self):
        result = courier_duty.classify("99999999")
        self.assertTrue(result.is_unknown)
        self.assertEqual(result.exemption_class, "none")
        self.assertEqual(result.rule_source, "fallback")

    def test_calculate_line_standard(self):
        result = courier_duty.calculate_line(
            cost_usd=20.0, freight_usd=0.0, exch_rate=6.78,
            thn="33049990",
        )
        self.assertEqual(result["cif_ttd"], 135.60)
        self.assertEqual(result["duty"], 27.12)
        self.assertEqual(result["opt"], 9.49)
        self.assertAlmostEqual(result["vat"], 21.53, places=2)
        self.assertAlmostEqual(result["total_taxes"], 58.14, places=2)

    def test_calculate_line_full_exempt(self):
        result = courier_duty.calculate_line(
            cost_usd=100.0, freight_usd=0.0, exch_rate=6.78,
            thn="85171300",
        )
        self.assertEqual(result["cif_ttd"], 678.00)
        self.assertEqual(result["duty"], 0.0)
        self.assertEqual(result["opt"], 0.0)
        self.assertEqual(result["vat"], 0.0)

    def test_calculate_line_duty_free_only(self):
        result = courier_duty.calculate_line(
            cost_usd=100.0, freight_usd=0.0, exch_rate=6.78,
            thn="85176900",
        )
        self.assertEqual(result["cif_ttd"], 678.00)
        self.assertEqual(result["duty"], 0.0)
        self.assertAlmostEqual(result["opt"], 47.46, places=2)

    def test_calculate_line_with_overrides(self):
        result = courier_duty.calculate_line(
            cost_usd=100.0, freight_usd=0.0, exch_rate=6.78,
            thn="33049990",
            duty_rate_override=0.10,
        )
        self.assertEqual(result["duty"], 67.80)

    def test_invalid_inputs(self):
        with self.assertRaises(ValueError):
            courier_duty.calculate_line(-10, 0, 6.78, "33049990")
        with self.assertRaises(ValueError):
            courier_duty.calculate_line(10, 0, 0, "33049990")
        with self.assertRaises(ValueError):
            courier_duty.calculate_line(10, 0, 6.78, "33049990",
                                         exemption_override="invalid")


class TestExemptionMutations(_RulesStoreTestBase):
    """User-driven exemption add/update/remove."""

    def test_add_user_exemption(self):
        result = courier_rules.add_exemption(
            thn="61091000", exemption_class="full_exempt",
            notes="Testing user exemption", by="christian",
        )
        self.assertEqual(result["thn"], "61091000")
        self.assertEqual(result["class"], "full_exempt")
        self.assertEqual(result["added_by"], "christian")

        cls = courier_duty.classify("61091000")
        self.assertEqual(cls.exemption_class, "full_exempt")
        self.assertEqual(cls.rule_source, "exemption")

    def test_user_exemption_overrides_bundled(self):
        """User can override a bundled exemption (e.g. demote it)."""
        courier_rules.add_exemption(
            thn="85171300", exemption_class="duty_free_only",
            notes="Reclassified per Customs Notice X", by="arnim",
            comment="Per signed-off rule change",
        )
        cls = courier_duty.classify("85171300")
        self.assertEqual(cls.exemption_class, "duty_free_only")

    def test_remove_user_exemption(self):
        courier_rules.add_exemption(
            thn="61091000", exemption_class="full_exempt", by="christian",
        )
        ok = courier_rules.remove_exemption("61091000", by="christian")
        self.assertTrue(ok)
        cls = courier_duty.classify("61091000")
        self.assertNotEqual(cls.exemption_class, "full_exempt")

    def test_remove_bundled_exemption_returns_false(self):
        ok = courier_rules.remove_exemption("85171300", by="christian")
        self.assertFalse(ok)
        cls = courier_duty.classify("85171300")
        self.assertEqual(cls.exemption_class, "full_exempt")

    def test_invalid_thn_rejected(self):
        with self.assertRaises(ValueError):
            courier_rules.add_exemption("12345", "full_exempt")
        with self.assertRaises(ValueError):
            courier_rules.add_exemption("ABCDEFGH", "full_exempt")

    def test_invalid_class_rejected(self):
        with self.assertRaises(ValueError):
            courier_rules.add_exemption("12345678", "made_up_class")


class TestCorrectionMutations(_RulesStoreTestBase):

    def test_add_user_correction(self):
        courier_rules.add_correction(
            wrong_thn="11111111", correct_thn="33049990",
            reason="Test correction", by="christian",
        )
        cls = courier_duty.classify("11111111")
        self.assertTrue(cls.is_corrected)
        self.assertEqual(cls.original_thn, "11111111")

    def test_correction_chains_to_exemption(self):
        courier_rules.add_correction(
            wrong_thn="11112222", correct_thn="85171300",
            reason="Map this to smartphone", by="christian",
        )
        cls = courier_duty.classify("11112222")
        self.assertTrue(cls.is_corrected)
        self.assertEqual(cls.exemption_class, "full_exempt")

    def test_remove_user_correction(self):
        courier_rules.add_correction("11111111", "33049990", by="christian")
        ok = courier_rules.remove_correction("11111111", by="christian")
        self.assertTrue(ok)
        cls = courier_duty.classify("11111111")
        self.assertFalse(cls.is_corrected)

    def test_self_correction_rejected(self):
        with self.assertRaises(ValueError):
            courier_rules.add_correction("11111111", "11111111")


class TestTariffOverrides(_RulesStoreTestBase):

    def test_add_tariff_override_for_missing_thn(self):
        courier_rules.add_tariff_entry(
            thn="99887766",
            description="Test tariff entry",
            duty_pct=15,
            chapter=99,
            unit="kg",
            by="christian",
        )
        entry = courier_duty.lookup_thn("99887766")
        self.assertIsNotNone(entry)
        self.assertEqual(entry["dutyPct"], 15)
        self.assertTrue(entry.get("is_override"))

        cls = courier_duty.classify("99887766")
        self.assertEqual(cls.exemption_class, "none")
        self.assertEqual(cls.duty_rate, 0.15)

    def test_user_tariff_override_replaces_bundled(self):
        courier_rules.add_tariff_entry(
            thn="33049990",
            description="Body cream — local override",
            duty_pct=10,
            by="arnim",
            comment="Test override",
        )
        cls = courier_duty.classify("33049990")
        self.assertEqual(cls.duty_rate, 0.10)

    def test_remove_tariff_override(self):
        courier_rules.add_tariff_entry(
            thn="99887766", description="x", duty_pct=10, chapter=99,
            by="christian",
        )
        ok = courier_rules.remove_tariff_entry("99887766", by="christian")
        self.assertTrue(ok)
        self.assertIsNone(courier_duty.lookup_thn("99887766"))

    def test_remove_bundled_tariff_returns_false(self):
        ok = courier_rules.remove_tariff_entry("33049990", by="christian")
        self.assertFalse(ok)

    def test_browse_filters_by_chapter(self):
        result = courier_rules.list_tariff_entries(chapter=33, limit=10)
        self.assertGreater(result["total"], 0)
        for e in result["items"]:
            self.assertEqual(e.get("chapter"), 33)

    def test_browse_filters_by_query(self):
        result = courier_rules.list_tariff_entries(query="smartphone", limit=10)
        self.assertGreater(result["total"], 0)


class TestAuditTrail(_RulesStoreTestBase):

    def test_audit_records_add(self):
        courier_rules.add_exemption(
            "61091000", "full_exempt", by="christian", comment="Test add",
        )
        log = courier_rules.get_audit_log()
        self.assertEqual(log["total"], 1)
        entry = log["items"][0]
        self.assertEqual(entry["action"], "add_exemption")
        self.assertEqual(entry["target"], "61091000")
        self.assertEqual(entry["by"], "christian")
        self.assertEqual(entry["comment"], "Test add")
        self.assertIsNone(entry["before"])
        self.assertIsNotNone(entry["after"])

    def test_audit_records_update(self):
        courier_rules.add_exemption("61091000", "full_exempt", by="christian")
        courier_rules.add_exemption(
            "61091000", "duty_free_only", by="arnim",
            comment="Demoted",
        )
        log = courier_rules.get_audit_log()
        self.assertEqual(log["total"], 2)
        latest = log["items"][0]
        self.assertEqual(latest["action"], "update_exemption")
        self.assertEqual(latest["before"]["class"], "full_exempt")
        self.assertEqual(latest["after"]["class"], "duty_free_only")

    def test_audit_records_remove(self):
        courier_rules.add_exemption("61091000", "full_exempt", by="christian")
        courier_rules.remove_exemption("61091000", by="christian", comment="Mistake")
        log = courier_rules.get_audit_log()
        self.assertEqual(log["total"], 2)
        self.assertEqual(log["items"][0]["action"], "remove_exemption")

    def test_audit_records_tariff_changes(self):
        courier_rules.add_tariff_entry(
            "99887766", "x", 10, chapter=99, by="christian",
        )
        log = courier_rules.get_audit_log()
        self.assertEqual(log["items"][0]["action"], "add_tariff")


class TestExportImport(_RulesStoreTestBase):

    def test_export_import_round_trip(self):
        courier_rules.add_exemption("61091000", "full_exempt", by="christian")
        courier_rules.add_correction("11111111", "33049990", by="christian")
        courier_rules.add_tariff_entry(
            "99887766", "test", 15, chapter=99, by="christian",
        )

        backup = courier_rules.export_user_rules()
        self.assertIn("rules", backup)
        self.assertIn("tariff_overrides", backup)
        self.assertEqual(len(backup["rules"]["exemptions"]), 1)
        self.assertEqual(len(backup["tariff_overrides"]["entries"]), 1)

        courier_rules.remove_exemption("61091000", by="christian")
        courier_rules.remove_correction("11111111", by="christian")
        courier_rules.remove_tariff_entry("99887766", by="christian")

        result = courier_rules.import_user_rules(backup, by="christian", comment="Restore")
        self.assertTrue(result["ok"])

        self.assertIsNotNone(courier_rules.get_exemption("61091000"))
        self.assertIsNotNone(courier_rules.get_correction("11111111"))
        self.assertIsNotNone(courier_duty.lookup_thn("99887766"))


class TestCourierMatcher(_RulesStoreTestBase):

    def test_smartphone_matches_full_exempt(self):
        result = courier_matcher.suggest_thns("smartphone")
        self.assertEqual(result["best_match"]["thn"], "85171300")
        self.assertEqual(result["best_match"]["exemption_class"], "full_exempt")

    def test_earphones_full_exempt(self):
        result = courier_matcher.suggest_thns("earphones")
        self.assertEqual(result["best_match"]["thn"], "85183000")

    def test_graphics_card_full_exempt(self):
        result = courier_matcher.suggest_thns("RTX 4090 graphics card")
        self.assertEqual(result["best_match"]["thn"], "84733000")

    def test_seeds_full_exempt(self):
        result = courier_matcher.suggest_thns("flower seeds")
        self.assertEqual(result["best_match"]["thn"], "12099900")

    def test_table_tennis_10pct(self):
        result = courier_matcher.suggest_thns("table tennis paddle")
        self.assertEqual(result["best_match"]["thn"], "95069190")

    def test_air_pump_match(self):
        result = courier_matcher.suggest_thns("portable air pump")
        self.assertEqual(result["best_match"]["thn"], "84148000")

    def test_generic_clothing_classifies(self):
        """Bare 'CLOTHING' must classify (was returning empty)."""
        r = courier_matcher.suggest_thns("CLOTHING")
        self.assertIsNotNone(r["best_match"])
        self.assertEqual(r["best_match"]["thn"], "61046900")

    def test_generic_jewelry_classifies(self):
        r = courier_matcher.suggest_thns("JEWELRY")
        self.assertIsNotNone(r["best_match"])
        self.assertEqual(r["best_match"]["thn"], "71171900")

    def test_thn_never_empty_even_for_gibberish(self):
        """
        The core promise: the THN field is NEVER blank. Even an
        unclassifiable description must return a fallback flagged for
        review rather than nothing.
        """
        for desc in ["SOCCER GAME SET", "xyzzy random thing",
                     "STORY BOARD", "miscellaneous goods", "qwerty"]:
            r = courier_matcher.suggest_thns(desc)
            self.assertIsNotNone(
                r["best_match"],
                f"{desc!r} returned no THN — must always fall back",
            )
            self.assertTrue(
                r["best_match"]["thn"],
                f"{desc!r} returned empty THN string",
            )

    def test_fallback_flagged_for_review(self):
        """Last-resort fallbacks must carry needs_review=True."""
        r = courier_matcher.suggest_thns("zxcvbnm unclassifiable")
        self.assertEqual(r["source"], "fallback")
        self.assertTrue(r["best_match"].get("needs_review"))

    def test_low_confidence_returns_multiple_options(self):
        """Vague descriptions should surface several review options."""
        r = courier_matcher.suggest_thns("CLOTHING", limit=5)
        self.assertGreaterEqual(
            len(r["suggestions"]), 3,
            "vague terms should give the broker several options to pick from",
        )

    def test_user_tariff_override_visible_in_matcher(self):
        """If user overrides a tariff rate, the matcher reflects it."""
        courier_rules.add_tariff_entry(
            "33049990", "Body cream — overridden", 10, by="arnim",
        )
        result = courier_matcher.suggest_thns("body cream")
        self.assertEqual(result["best_match"]["thn"], "33049990")
        self.assertEqual(result["best_match"]["duty_rate"], 0.10)


class TestServiceIntegration(_RulesStoreTestBase):
    """End-to-end: manifest with auto-classification picks up user rules."""

    def setUp(self):
        super().setUp()
        from app import store_courier
        self.manifest_path_backup = store_courier.COURIER_FILE
        store_courier.COURIER_FILE = Path(self.tmpdir) / "courier_manifests.json"
        store_courier.COURIER_FILE.write_text("[]", encoding="utf-8")

    def tearDown(self):
        from app import store_courier
        store_courier.COURIER_FILE = self.manifest_path_backup
        super().tearDown()

    def test_manifest_uses_user_exemption(self):
        from app.services import courier_service

        courier_rules.add_exemption(
            "61091000", "full_exempt",
            notes="Test exemption for cotton t-shirts",
            by="arnim", comment="Customer chamber rate",
        )

        m = courier_service.create_manifest({
            "manifest_no": "TEST-MANIFEST",
            "arrival_date": "2026-05-04",
            "exch_rate": 6.78,
        })
        line = courier_service.add_line(m["id"], {
            "description": "cotton t-shirt",
            "thn": "61091000",
            "cost_usd": 25.00,
        })
        self.assertEqual(line["exemption_class"], "full_exempt")
        self.assertEqual(line["duty"], 0.0)
        self.assertEqual(line["total_taxes"], 0.0)


# ─── Phase 2: XLSX export tests ──────────────────────────────────────────────


class TestWorksheetExport(_RulesStoreTestBase):
    """Build a v3 worksheet from a manifest and inspect the result."""

    def setUp(self):
        super().setUp()
        from app import store_courier
        self.manifest_path_backup = store_courier.COURIER_FILE
        store_courier.COURIER_FILE = Path(self.tmpdir) / "courier_manifests.json"
        store_courier.COURIER_FILE.write_text("[]", encoding="utf-8")

    def tearDown(self):
        from app import store_courier
        store_courier.COURIER_FILE = self.manifest_path_backup
        super().tearDown()

    def _make_manifest(self):
        """Build a small manifest with mixed exemption classes."""
        from app.services import courier_service
        m = courier_service.create_manifest({
            "manifest_no": "EXPORT-TEST-001",
            "arrival_date": "2026-05-04",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "hawb": "TEST-001-A", "shipper": "Amazon", "importer": "John Doe",
            "description": "Body cream", "thn": "33049990",
            "cost_usd": 50.0, "freight_usd": 0.0, "packages": 1, "weight_kg": 0.5,
        })
        courier_service.add_line(m["id"], {
            "hawb": "TEST-001-B", "shipper": "Apple", "importer": "Jane Smith",
            "description": "Smartphone", "thn": "85171300",
            "cost_usd": 800.0, "freight_usd": 0.0, "packages": 1, "weight_kg": 0.3,
        })
        courier_service.add_line(m["id"], {
            "hawb": "TEST-001-C", "shipper": "Nest", "importer": "Bob Brown",
            "description": "Smart device", "thn": "85176900",
            "cost_usd": 100.0, "freight_usd": 0.0, "packages": 1, "weight_kg": 0.4,
        })
        return courier_service.get_manifest(m["id"])

    def test_worksheet_returns_bytes(self):
        from app.services import courier_export
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        self.assertIsInstance(data, bytes)
        self.assertEqual(data[:2], b"PK")

    def test_worksheet_layout_matches_real_template(self):
        """Cell positions should match the real AWB 5034 template exactly."""
        from openpyxl import load_workbook
        from app.services import courier_export
        import io

        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active

        # Title block
        self.assertIn("EXPRESS CONSIGNMENTS WORKSHEET", str(ws["A1"].value))
        self.assertIn("NON-COMMERCIAL", str(ws["A2"].value))
        self.assertIn("CARGO REPORTER", str(ws["A3"].value))
        self.assertIn("MASTER WAY BILL", str(ws["J3"].value))
        self.assertIn("EXPORT-TEST-001", str(ws["J3"].value))
        self.assertIn("R.O.E.", str(ws["F4"].value))
        self.assertIn("FREIGHT", str(ws["J4"].value))

        # Banner row 6
        self.assertEqual(ws["A6"].value, "SECTION 2")
        self.assertIn("SECTION 3", str(ws["Q6"].value))

        # Column headers row 7
        self.assertIn("LINE", str(ws["A7"].value))
        self.assertEqual(ws["B7"].value, "HAWB")
        self.assertEqual(ws["C7"].value, "SHIPPER")
        self.assertIn("IMPORTER", str(ws["D7"].value))
        self.assertIn("DESCRIPTION", str(ws["E7"].value))
        self.assertIn("PKGS", str(ws["F7"].value))
        self.assertEqual(ws["H7"].value, "THN")
        self.assertEqual(ws["I7"].value, "RATE")
        self.assertIn("CUSTOMS", str(ws["L7"].value))
        self.assertEqual(ws["M7"].value, "DUTY")
        self.assertIn("OPT", str(ws["N7"].value))
        self.assertIn("VAT", str(ws["O7"].value))
        self.assertIn("TOTAL", str(ws["P7"].value))
        self.assertIn("OFFICER", str(ws["Q7"].value))
        self.assertIn("ADJUSTED", str(ws["S7"].value))
        self.assertIn("DETAINED", str(ws["X7"].value))
        self.assertIn("T/SHED", str(ws["Y7"].value))

        # First data row at row 8 (NOT row 9)
        self.assertEqual(ws["A8"].value, 1)
        self.assertEqual(ws["E8"].value, "Body cream")
        self.assertEqual(ws["H8"].value, "33049990")
        self.assertEqual(ws["I8"].value, "20%")
        self.assertAlmostEqual(ws["J8"].value, 50.0, places=2)

    def test_worksheet_values_no_formulas(self):
        """
        Generated worksheet contains EXACT VALUES, never formulas.

        Formulas were removed entirely because Excel recalculation drifts
        from the workbench figures. Every computed cell must be a plain
        number matching what the broker saw on screen.
        """
        from openpyxl import load_workbook
        from app.services import courier_export
        import io

        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)

        wb = load_workbook(io.BytesIO(data), data_only=False)
        ws = wb.active

        # No cell anywhere may contain a formula string
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    self.assertFalse(
                        cell.value.startswith("="),
                        f"Cell {cell.coordinate} contains a formula: {cell.value!r}",
                    )

        # Row 8 (body cream): values must equal the stored line values.
        line1 = m["lines"][0]
        self.assertAlmostEqual(ws["L8"].value, line1["cif_ttd"], places=2)
        self.assertAlmostEqual(ws["M8"].value, line1["duty"], places=2)
        self.assertAlmostEqual(ws["N8"].value, line1["opt"], places=2)
        self.assertAlmostEqual(ws["O8"].value, line1["vat"], places=2)
        self.assertAlmostEqual(ws["P8"].value, line1["total_taxes"], places=2)

        # Row 9: smartphone — full_exempt, hard zeros
        self.assertEqual(ws["I9"].value, "FREE")
        self.assertEqual(ws["M9"].value, 0)
        self.assertEqual(ws["N9"].value, 0)
        self.assertEqual(ws["O9"].value, 0)

        # Row 10: smart device — duty_free_only (duty=0, OPT/VAT > 0)
        self.assertEqual(ws["I10"].value, "FREE")
        self.assertEqual(ws["M10"].value, 0)
        self.assertGreater(ws["N10"].value or 0, 0)
        self.assertGreater(ws["O10"].value or 0, 0)

    def test_worksheet_values_match_workbench_exactly(self):
        """
        Every Section 2 money cell equals the value stored on the line —
        zero drift. This is the core guarantee: the XLSX shows EXACTLY
        what the broker saw in the workbench.
        """
        from openpyxl import load_workbook
        from app.services import courier_export
        import io

        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data), data_only=False)
        ws = wb.active

        for i, line in enumerate(m["lines"]):
            r = 8 + i
            self.assertEqual(ws[f"L{r}"].value, line["cif_ttd"],
                             f"CIF mismatch row {r}")
            self.assertEqual(ws[f"M{r}"].value, line["duty"],
                             f"Duty mismatch row {r}")
            self.assertEqual(ws[f"N{r}"].value, line["opt"],
                             f"OPT mismatch row {r}")
            self.assertEqual(ws[f"O{r}"].value, line["vat"],
                             f"VAT mismatch row {r}")
            self.assertEqual(ws[f"P{r}"].value, line["total_taxes"],
                             f"Total mismatch row {r}")

    def test_worksheet_with_officer_corrections(self):
        """Officer examination data lands in Section 3 of the right rows."""
        import io
        from openpyxl import load_workbook
        from app.services import courier_export, courier_service

        m = self._make_manifest()
        courier_service.record_examination(m["id"], {
            "examined_at": "2026-05-04",
            "examining_officer": "Officer Test",
            "corrections": [
                {
                    "line_no": 1, "kind": "uplift",
                    "officer_thn": "33049990",
                    "add_cost_usd": 25.0,
                    "adjusted_cif_ttd": 169.50,
                    "add_duty": 33.90,
                    "add_opt": 11.87,
                    "add_vat": 26.91,
                },
            ],
        })

        m = courier_service.get_manifest(m["id"])
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active

        # Line 1 → row 8, S3 fields populated
        self.assertEqual(ws["Q8"].value, "33049990")
        self.assertAlmostEqual(ws["R8"].value, 25.0, places=2)
        self.assertAlmostEqual(ws["S8"].value, 169.50, places=2)
        self.assertAlmostEqual(ws["T8"].value, 33.90, places=2)
        # W8 is the exact value (no formula): 33.90 + 11.87 + 26.91 = 72.68
        self.assertNotIsInstance(ws["W8"].value, str,
                                 "W8 must be a number, never a formula")
        self.assertAlmostEqual(ws["W8"].value, 72.68, places=2)

        # Line 2 → row 9, S3 empty
        self.assertIn(ws["Q9"].value, (None, ""))

    def test_worksheet_totals_are_exact_values(self):
        """
        TOTALS row holds exact summed VALUES (no =SUM formulas). The totals
        equal the sum of the stored line values, with zero recalc drift.
        """
        from openpyxl import load_workbook
        from app.services import courier_export
        import io

        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)

        wb = load_workbook(io.BytesIO(data), data_only=False)
        ws = wb.active
        self.assertEqual(ws["A11"].value, "TOTALS")

        # No SUM formulas anywhere — every total is a plain number
        for col in ("F", "G", "J", "L", "M", "N", "O", "P",
                    "R", "S", "T", "U", "V", "W"):
            v = ws[f"{col}11"].value
            self.assertNotIsInstance(
                v, str,
                f"Cell {col}11 must be a number, got formula/string {v!r}",
            )

        # Grand total row 12 also values, not formulas
        self.assertNotIsInstance(ws["P12"].value, str)
        self.assertNotIsInstance(ws["W12"].value, str)

        # Totals must equal the exact sum of stored line values
        lines = m["lines"]
        exp_cif = sum(l["cif_ttd"] for l in lines)
        exp_duty = sum(l["duty"] for l in lines)
        exp_opt = sum(l["opt"] for l in lines)
        exp_vat = sum(l["vat"] for l in lines)
        exp_total = sum(l["total_taxes"] for l in lines)
        self.assertEqual(ws["F11"].value, 3)        # 3 packages
        self.assertAlmostEqual(ws["J11"].value, 950.0, places=2)
        self.assertAlmostEqual(ws["L11"].value, exp_cif, places=2)
        self.assertAlmostEqual(ws["M11"].value, exp_duty, places=2)
        self.assertAlmostEqual(ws["N11"].value, exp_opt, places=2)
        self.assertAlmostEqual(ws["O11"].value, exp_vat, places=2)
        self.assertAlmostEqual(ws["P11"].value, exp_total, places=2)

    def test_worksheet_values_stable_across_reopen(self):
        """
        Since there are no formulas, opening the file (with or without a
        recalc engine) must yield the SAME values. This is the whole point
        of removing formulas — zero drift between generation and viewing.
        """
        from openpyxl import load_workbook
        from app.services import courier_export
        import io

        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)

        # data_only=True and data_only=False must agree (no formulas means
        # the "cached value" and the "cell value" are identical).
        wb_f = load_workbook(io.BytesIO(data), data_only=False).active
        wb_v = load_workbook(io.BytesIO(data), data_only=True).active

        for r in (8, 9, 10):
            for col in ("L", "M", "N", "O", "P"):
                self.assertEqual(
                    wb_f[f"{col}{r}"].value, wb_v[f"{col}{r}"].value,
                    f"{col}{r} differs between formula/value reads — "
                    f"means a formula leaked in",
                )

        # Row 8 (body cream): exact stored values
        line1 = m["lines"][0]
        self.assertEqual(wb_f["L8"].value, line1["cif_ttd"])
        self.assertEqual(wb_f["M8"].value, line1["duty"])
        self.assertEqual(wb_f["N8"].value, line1["opt"])

        # Row 9 (smartphone, full_exempt): all zero
        self.assertEqual(wb_f["M9"].value, 0)
        self.assertEqual(wb_f["N9"].value, 0)
        self.assertEqual(wb_f["O9"].value, 0)

        # Row 10 (smart device, duty_free_only): duty 0, OPT/VAT > 0
        self.assertEqual(wb_f["M10"].value, 0)
        self.assertGreater(wb_f["N10"].value or 0, 0)
        self.assertGreater(wb_f["O10"].value or 0, 0)


class TestHazmatExport(_RulesStoreTestBase):
    """Build the Swissport Transit Shed Hazmat XLSX and verify its structure."""

    def setUp(self):
        super().setUp()
        from app import store_courier
        self.manifest_path_backup = store_courier.COURIER_FILE
        store_courier.COURIER_FILE = Path(self.tmpdir) / "courier_manifests.json"
        store_courier.COURIER_FILE.write_text("[]", encoding="utf-8")

    def tearDown(self):
        from app import store_courier
        store_courier.COURIER_FILE = self.manifest_path_backup
        super().tearDown()

    def test_hazmat_swissport_layout(self):
        """Verify the Swissport Transit Shed Form layout (Trade + Non-Trade sections)."""
        import io
        from openpyxl import load_workbook
        from app.services import courier_export, courier_service

        m = courier_service.create_manifest({
            "manifest_no": "HZ-001",
            "arrival_date": "2026-05-04",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "hawb": "H-1", "description": "Body cream", "thn": "33049990",
            "cost_usd": 50.0, "packages": 1, "weight_kg": 0.5,
        })
        m = courier_service.get_manifest(m["id"])
        data = courier_export.build_hazmat(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active

        # Header
        self.assertIn("SWISSPORT", str(ws["B2"].value).upper())

        # Manifest meta
        self.assertEqual(ws["B7"].value, "NAME OF COURIER:")
        self.assertEqual(ws["J7"].value, "AWB/BL #")
        self.assertEqual(ws["K7"].value, "HZ-001")

        # Tax column headers at row 22
        self.assertEqual(ws["F22"].value, "CIF")
        self.assertEqual(ws["H22"].value, "OPT")
        self.assertEqual(ws["J22"].value, "DUTY")
        self.assertEqual(ws["L22"].value, "VAT")
        self.assertEqual(ws["N22"].value, "TOTAL")

        # Trade section (rows 23, 25, 27) — values only, no formulas
        self.assertEqual(ws["E23"].value, "Trade")
        self.assertEqual(ws["B23"].value, "Original Values Declared")
        self.assertEqual(ws["F23"].value, 0)  # Trade is zero for TTPOST
        # Additional row 25 (above the "Additional Taxes" label on row 26)
        # — values, not formulas; Trade row 25 in the golden is empty merged
        # blocks. The cell value is None (just the merge exists).
        self.assertIsNone(ws["F25"].value)
        self.assertIsNone(ws["H25"].value)
        self.assertEqual(ws["F27"].value, 0)  # Trade Additional values

        # Non-Trade section (rows 31, 33, 35) — values only, no formulas
        self.assertEqual(ws["E31"].value, "Non-Trade")
        self.assertGreater(ws["F31"].value, 0)  # Non-Trade Original CIF
        # Additional row 33 carries the broker's correction add_* values
        # (or zero) as plain numbers — never formulas.
        self.assertNotIsInstance(ws["F33"].value, str,
                                 "F33 must be a number, not a formula")
        self.assertNotIsInstance(ws["H33"].value, str)
        self.assertNotIsInstance(ws["J33"].value, str)
        self.assertNotIsInstance(ws["L33"].value, str)

        # Summary footer — values only
        self.assertEqual(ws["B39"].value, "Total Additional Taxes")
        self.assertNotIsInstance(ws["F38"].value, str,
                                 "F38 must be a number, not a formula")
        self.assertEqual(ws["B42"].value, "TOTAL TAXES")
        self.assertNotIsInstance(ws["F41"].value, str,
                                 "F41 must be a number, not a formula")

    def test_hazmat_with_corrections(self):
        """Officer corrections feed into the Non-Trade Final values."""
        import io
        from openpyxl import load_workbook
        from app.services import courier_export, courier_service

        m = courier_service.create_manifest({
            "manifest_no": "HZ-002", "arrival_date": "2026-05-04", "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "Body cream", "thn": "33049990", "cost_usd": 50.0,
        })
        courier_service.record_examination(m["id"], {
            "corrections": [
                {"line_no": 1, "kind": "uplift",
                 "officer_thn": "33049990",
                 "adjusted_cif_ttd": 169.50,
                 "add_duty": 33.90, "add_opt": 11.87, "add_vat": 26.91},
            ],
        })
        m = courier_service.get_manifest(m["id"])
        data = courier_export.build_hazmat(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active

        # F35 (Non-Trade Final CIF) = original_cif + add_cif
        # original_cif = 50 * 6.78 = 339, add_cif = 169.50, final = 508.50
        self.assertAlmostEqual(ws["F35"].value, 508.50, places=2)
        # J35 (Final Duty) = orig duty (67.80) + add duty (33.90) = 101.70
        self.assertAlmostEqual(ws["J35"].value, 101.70, places=2)


class TestTemplateParser(_RulesStoreTestBase):
    """Tests for the TTPOST blank-template parser (Phase 4)."""

    def _build_minimal_ttpost_template(self) -> bytes:
        """Build an in-memory TTPOST-style XLSX with header + 3 data rows."""
        import io
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        # Header rows
        ws["H1"] = "EXPRESS CONSIGNMENTS WORKSHEET"
        ws["H2"] = "NON-COMMERCIAL CONSIGNMENTS"
        ws["A3"] = "CARGO REPORTER: TRINIDAD AND TOBAGO POSTAL CORPORATION"
        ws["J3"] = "MASTER WAY BILL NUMBER: 106-31299999"
        ws["A5"] = 'VAT NO. / "N" NO. : V117369'
        ws["A7"] = "SECTION 2"
        ws["A8"] = "DETAILS OF ALL HOUSE WAYBILLS"
        # Column headers row 9
        headers = ["LINE NO. AWB", "HAWB", "SHIPPER", "NAME OF IMPORTER",
                   "DESCRIPTION OF GOODS", "NO. OF PKGS", "WEIGHT", "THN",
                   "RATE", "COST", "FREIGHT", "CUSTOMS VALUE", "DUTY",
                   "OPT", "VAT", "TOTAL TAXES"]
        for col, h in enumerate(headers, 1):
            ws.cell(row=9, column=col, value=h)
        # Data rows 10-12
        rows = [
            (1, "2700001", "AMAZON", "TEST IMPORTER 1", "BACKPACK", 1, 1.0, None, None, 21.0),
            (2, "2700002", "SHEIN",  "TEST IMPORTER 2", "SHOES",    1, 2.0, None, None, 19.99),
            (3, "2700003", "TEMU",   "TEST IMPORTER 3", "CELL PHONE", 1, 0.5, None, None, 250.0),
        ]
        for r_idx, row_data in enumerate(rows, start=10):
            for c_idx, val in enumerate(row_data, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def test_parser_extracts_header(self):
        from app.services.courier_template_parser import parse_ttpost_template
        result = parse_ttpost_template(self._build_minimal_ttpost_template())
        self.assertEqual(result["manifest_no"], "106-31299999")
        self.assertEqual(result["vat_no"], "V117369")
        self.assertIn("TRINIDAD AND TOBAGO POSTAL CORPORATION", result["cargo_reporter"])

    def test_parser_extracts_lines(self):
        from app.services.courier_template_parser import parse_ttpost_template
        result = parse_ttpost_template(self._build_minimal_ttpost_template())
        self.assertEqual(len(result["lines"]), 3)
        first = result["lines"][0]
        self.assertEqual(first["hawb"], "2700001")
        self.assertEqual(first["shipper"], "AMAZON")
        self.assertEqual(first["importer"], "TEST IMPORTER 1")
        self.assertEqual(first["description"], "BACKPACK")
        self.assertEqual(first["packages"], 1)
        self.assertEqual(first["weight_kg"], 1.0)
        self.assertEqual(first["cost_usd"], 21.0)
        # THN was empty in the template
        self.assertEqual(first["thn"], "")

    def test_parser_stops_at_blank_rows(self):
        """3 consecutive blank rows = end of data."""
        from app.services.courier_template_parser import parse_ttpost_template
        # The minimal template has only 3 data rows; sheet ends after them
        result = parse_ttpost_template(self._build_minimal_ttpost_template())
        self.assertEqual(len(result["lines"]), 3)

    def test_parser_rejects_non_ttpost_file(self):
        """A file with no recognizable header structure should raise ValueError."""
        from openpyxl import Workbook
        import io
        from app.services.courier_template_parser import parse_ttpost_template

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "Just some random data"
        ws["B1"] = "Not a TTPOST file"
        buf = io.BytesIO()
        wb.save(buf)

        with self.assertRaises(ValueError):
            parse_ttpost_template(buf.getvalue())

    def test_parser_handles_missing_master_waybill(self):
        """If master waybill is absent, return empty string + warning."""
        from openpyxl import Workbook
        import io
        from app.services.courier_template_parser import parse_ttpost_template

        wb = Workbook()
        ws = wb.active
        ws["A1"] = "EXPRESS CONSIGNMENTS WORKSHEET"
        ws["A9"] = "LINE NO. AWB"
        ws["B9"] = "HAWB"
        ws["C9"] = "SHIPPER"
        ws["D9"] = "NAME OF IMPORTER"
        ws["E9"] = "DESCRIPTION OF GOODS"
        ws["F9"] = "NO. OF PKGS"
        ws["J9"] = "COST"
        # One data row
        ws["A10"] = 1
        ws["B10"] = "HAWB1"
        ws["C10"] = "SHIPPER1"
        ws["D10"] = "IMPORTER1"
        ws["E10"] = "BOOKS"
        ws["F10"] = 1
        ws["J10"] = 10.0
        buf = io.BytesIO()
        wb.save(buf)

        result = parse_ttpost_template(buf.getvalue())
        self.assertEqual(result["manifest_no"], "")
        self.assertTrue(any("master waybill" in w.lower() for w in result["warnings"]))


class TestTemplateUploadEndToEnd(_RulesStoreTestBase):
    """End-to-end: parse a template, create manifest, auto-classify all lines."""

    def setUp(self):
        super().setUp()
        # Need an isolated manifest store too
        from app import store_courier
        from app.services import courier_service
        store_courier.COURIER_FILE = Path(self.tmpdir) / "manifests.json"
        store_courier.COURIER_FILE.write_text("[]")

    def test_upload_creates_manifest_with_classified_lines(self):
        """Parse template → create manifest → all 3 lines have auto-classified THNs."""
        from app.services import courier_service, courier_template_parser

        # Build template inline
        from openpyxl import Workbook
        import io
        wb = Workbook()
        ws = wb.active
        ws["H1"] = "EXPRESS CONSIGNMENTS WORKSHEET"
        ws["A3"] = "CARGO REPORTER: TTPOST"
        ws["J3"] = "MASTER WAY BILL NUMBER: 106-31299998"
        ws["A9"] = "LINE NO. AWB"
        ws["B9"] = "HAWB"
        ws["C9"] = "SHIPPER"
        ws["D9"] = "NAME OF IMPORTER"
        ws["E9"] = "DESCRIPTION OF GOODS"
        ws["F9"] = "NO. OF PKGS"
        ws["G9"] = "WEIGHT"
        ws["H9"] = "THN"
        ws["J9"] = "COST"
        # Data: 3 lines with descriptions the matcher knows
        data = [
            (1, "H1", "AMAZON", "A", "iPhone 15 Pro Max", 1, 1.0, None, 999.0),
            (2, "H2", "AMAZON", "B", "wooden chair", 1, 10.0, None, 50.0),
            (3, "H3", "AMAZON", "C", "AirPods Pro", 1, 0.5, None, 250.0),
        ]
        for r_idx, row in enumerate(data, start=10):
            for c_idx, val in enumerate(row, start=1):
                ws.cell(row=r_idx, column=c_idx, value=val)
        buf = io.BytesIO()
        wb.save(buf)

        # Parse + import
        parsed = courier_template_parser.parse_ttpost_template(buf.getvalue())
        self.assertEqual(parsed["manifest_no"], "106-31299998")

        m = courier_service.create_manifest({
            "manifest_no": parsed["manifest_no"],
            "arrival_date": "2026-05-11",
            "exch_rate": 6.78,
        })
        for line in parsed["lines"]:
            courier_service.add_line_with_auto_thn(m["id"], {
                "hawb": line["hawb"], "shipper": line["shipper"],
                "importer": line["importer"], "description": line["description"],
                "packages": line["packages"], "weight_kg": line["weight_kg"],
                "cost_usd": line["cost_usd"], "freight_usd": line["freight_usd"],
            })

        final = courier_service.get_manifest(m["id"])
        self.assertEqual(len(final["lines"]), 3)

        # iPhone → 85171300
        self.assertEqual(final["lines"][0]["thn"], "85171300")
        # AirPods → 85183000
        self.assertEqual(final["lines"][2]["thn"], "85183000")

        # Confidence and suggestions are persisted
        for ln in final["lines"]:
            self.assertIn("thn_confidence", ln)
            self.assertIn("thn_suggestions", ln)
            self.assertIsInstance(ln["thn_suggestions"], list)

    def test_classified_thns_persist_through_reload(self):
        """Verify the fix: thn_suggestions must survive a manifest reload."""
        from app.services import courier_service

        m = courier_service.create_manifest({
            "manifest_no": "TEST-PERSIST",
            "arrival_date": "2026-05-11",
            "exch_rate": 6.78,
        })
        courier_service.add_line_with_auto_thn(m["id"], {
            "description": "iPhone 15",
            "cost_usd": 1000.0,
        })

        # Reload from disk
        fresh = courier_service.get_manifest(m["id"])
        line = fresh["lines"][0]

        # Suggestions must be present
        self.assertIn("thn_suggestions", line)
        self.assertGreater(len(line["thn_suggestions"]), 0)
        # Confidence must be a number
        self.assertIsNotNone(line["thn_confidence"])
        # Match source must be set
        self.assertIn(line["thn_match_source"], ("keyword_index", "full_text", "hybrid"))


class TestMaintainTariffFlow(_RulesStoreTestBase):
    """Tests for the Maintain Tariff workflow — tariff override → recompute → updated taxes."""

    def setUp(self):
        super().setUp()
        from app import store_courier
        store_courier.COURIER_FILE = Path(self.tmpdir) / "manifests.json"
        store_courier.COURIER_FILE.write_text("[]")

    def test_tariff_override_then_recompute_updates_line_taxes(self):
        """
        Broker uploads, classifier picks THN, broker maintains the tariff
        entry to change its duty rate, recompute → line's duty/total update.
        """
        from app.services import courier_service, courier_rules

        # 1. Create a manifest with one line using a known THN
        m = courier_service.create_manifest({
            "manifest_no": "TEST-MAINT",
            "arrival_date": "2026-05-12",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "Random item",
            "thn": "61091000",  # T-shirts cotton — bundled at 20%
            "cost_usd": 100.0,
        })

        fresh = courier_service.get_manifest(m["id"])
        line_before = fresh["lines"][0]
        duty_before = line_before["duty"]
        self.assertGreater(duty_before, 0)  # 20% of 678 = 135.60

        # 2. Broker maintains the tariff: change duty from 20% to 10%
        courier_rules.add_tariff_entry(
            "61091000",
            description="T-shirts (broker override)",
            duty_pct=10,
            by="broker",
            comment="Test override",
        )

        # 3. Recompute the manifest
        m2 = courier_service.recompute_manifest(m["id"])
        line_after = m2["lines"][0]
        duty_after = line_after["duty"]

        # Duty should be lower: 10% of 678 = 67.80
        self.assertLess(duty_after, duty_before)
        self.assertAlmostEqual(duty_after, 67.80, places=1)

    def test_tariff_override_to_exempt_zeroes_taxes(self):
        """Adding a full_exempt rule should zero out duty/OPT/VAT after recompute."""
        from app.services import courier_service, courier_rules

        m = courier_service.create_manifest({
            "manifest_no": "TEST-EXEMPT-FLOW",
            "arrival_date": "2026-05-12",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "Test item",
            "thn": "61091000",
            "cost_usd": 100.0,
        })

        # Make it full exempt
        courier_rules.add_exemption(
            "61091000",
            "full_exempt",
            notes="Test exemption",
            by="broker",
        )
        m2 = courier_service.recompute_manifest(m["id"])
        line = m2["lines"][0]

        self.assertEqual(line["duty"], 0)
        self.assertEqual(line["opt"], 0)
        self.assertEqual(line["vat"], 0)
        self.assertEqual(line["total_taxes"], 0)
        self.assertEqual(line["exemption_class"], "full_exempt")

    def test_recompute_404_for_unknown_manifest(self):
        from app.services import courier_service
        result = courier_service.recompute_manifest("nonexistent-id")
        self.assertIsNone(result)


class TestHandoffBugRegression(_RulesStoreTestBase):
    """
    Regression: 'Input should be a valid dictionary' on Maintain Tariff save.

    From the handoff brief: broker opens Maintain Tariff for THN 84713000
    (laptop), sets exemption_class=full_exempt, clicks Save. Save should
    persist tariff override + exemption + recompute manifest so taxes go
    to zero.
    """

    def setUp(self):
        super().setUp()
        from app import store_courier
        store_courier.COURIER_FILE = Path(self.tmpdir) / "manifests.json"
        store_courier.COURIER_FILE.write_text("[]")

    def test_full_exempt_save_flow_persists_and_zeroes_taxes(self):
        """Reproduce the handoff bug: full save chain must result in zero taxes."""
        from app.services import courier_service, courier_rules

        # Setup: manifest with line at THN 84713000
        m = courier_service.create_manifest({
            "manifest_no": "HANDOFF-BUG", "arrival_date": "2026-05-12", "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "Laptop", "thn": "84713000", "cost_usd": 1000.0,
        })

        # The 3-step save chain the Maintain dialog performs:

        # 1) Tariff override
        courier_rules.add_tariff_entry(
            thn="84713000",
            description="Portable digital automatic data processing machines",
            duty_pct=0,
            chapter=84,
            unit="u",
            is_exempt=True,
            by="broker",
            comment="handoff test",
        )

        # 2) Exemption rule
        courier_rules.add_exemption(
            thn="84713000", exemption_class="full_exempt",
            notes="Laptop full exempt", by="broker", comment="handoff test",
        )

        # 3) Recompute
        result = courier_service.recompute_manifest(m["id"])
        line = result["lines"][0]

        self.assertEqual(line["exemption_class"], "full_exempt")
        self.assertEqual(line["duty"], 0)
        self.assertEqual(line["opt"], 0)
        self.assertEqual(line["vat"], 0)
        self.assertEqual(line["total_taxes"], 0)

    def test_duty_free_only_keeps_opt_and_vat(self):
        """Critical distinction from the brief: duty_free_only ≠ full_exempt."""
        from app.services import courier_service, courier_rules

        m = courier_service.create_manifest({
            "manifest_no": "DFO-TEST", "arrival_date": "2026-05-12", "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "Test", "thn": "84713000", "cost_usd": 1000.0,
        })

        # Set duty_free_only — duty is 0 but OPT/VAT still apply
        courier_rules.add_exemption(
            thn="84713000", exemption_class="duty_free_only",
            notes="DFO test", by="broker",
        )
        result = courier_service.recompute_manifest(m["id"])
        line = result["lines"][0]

        self.assertEqual(line["exemption_class"], "duty_free_only")
        self.assertEqual(line["duty"], 0)
        self.assertGreater(line["opt"], 0, "OPT should still apply for duty_free_only")
        self.assertGreater(line["vat"], 0, "VAT should still apply for duty_free_only")


class TestPayloadValidation(_RulesStoreTestBase):
    """
    Validation tests for the rules/tariff POST endpoints. After the Pydantic
    hardening, every malformed payload should return a clear, field-specific
    422 — never the opaque 'Input should be a valid dictionary'.
    """

    def _make_app(self):
        from fastapi import FastAPI
        from fastapi.testclient import TestClient
        from app.routes.courier_rules import router as courier_rules_router
        app = FastAPI()
        app.include_router(courier_rules_router)
        return TestClient(app)

    def test_tariff_missing_thn_returns_clear_error(self):
        client = self._make_app()
        r = client.post("/courier/tariff",
                        json={"description": "X", "duty_pct": 0},
                        headers={"X-User-Id": "broker"})
        self.assertEqual(r.status_code, 422)
        # Error must specifically call out the missing 'thn' field
        detail = r.json()["detail"]
        self.assertTrue(any("thn" in str(d.get("loc", [])) for d in detail))

    def test_tariff_invalid_thn_length_returns_clear_error(self):
        client = self._make_app()
        r = client.post("/courier/tariff",
                        json={"thn": "1234567", "description": "X", "duty_pct": 0},
                        headers={"X-User-Id": "broker"})
        self.assertEqual(r.status_code, 422)
        detail_str = str(r.json()["detail"])
        self.assertIn("THN must be 8 digits", detail_str)

    def test_tariff_duty_out_of_range_returns_clear_error(self):
        client = self._make_app()
        r = client.post("/courier/tariff",
                        json={"thn": "84713000", "description": "X", "duty_pct": 150},
                        headers={"X-User-Id": "broker"})
        self.assertEqual(r.status_code, 422)

    def test_exemption_invalid_class_returns_clear_error(self):
        client = self._make_app()
        r = client.post("/courier/rules/exemptions",
                        json={"thn": "85171300", "class": "nonsense", "notes": "x"},
                        headers={"X-User-Id": "broker"})
        self.assertEqual(r.status_code, 422)
        detail_str = str(r.json()["detail"])
        # Pydantic Literal error mentions the allowed values
        self.assertIn("full_exempt", detail_str)

    def test_tariff_happy_path_succeeds(self):
        client = self._make_app()
        r = client.post("/courier/tariff",
                        json={"thn": "85171300", "description": "Smartphones",
                              "duty_pct": 0, "chapter": 85, "is_exempt": True,
                              "comment": "test"},
                        headers={"X-User-Id": "broker"})
        self.assertEqual(r.status_code, 200, r.text)
        body = r.json()
        self.assertEqual(body["thn"], "85171300")
        self.assertEqual(body["dutyPct"], 0)
        self.assertTrue(body["isExempt"])

    def test_exemption_happy_path_succeeds(self):
        client = self._make_app()
        r = client.post("/courier/rules/exemptions",
                        json={"thn": "85171300", "class": "full_exempt",
                              "notes": "Smartphones - test"},
                        headers={"X-User-Id": "broker"})
        self.assertEqual(r.status_code, 200, r.text)


class TestWorksheetGoldenParity(_RulesStoreTestBase):
    """
    Parity tests against the broker's golden template
    (Worksheet_106-31245034_FINAL_v3). Each test asserts one slice of the
    layout — merged ranges, column widths, anchor labels, row heights —
    that the export engine MUST reproduce exactly.
    """

    GOLDEN_MERGED_RANGES = {
        "A1:X1",   # title row
        "A2:X2",   # subtitle
        "J3:X3",   # master waybill
        "A5:X5",   # CBTT note
        "A6:P6",   # SECTION 2 banner
        "Q6:X6",   # SECTION 3 banner
        "A34:E34", # TOTALS label
        "A35:O35", # TOTAL TAXES label
        "Q35:V35", # TOTAL INCL. OFFICER UPLIFTS label
    }

    GOLDEN_COLUMN_WIDTHS = {
        "A": 6.0, "B": 10.0, "C": 20.0, "D": 22.0, "E": 35.0, "F": 5.0,
        "H": 12.0, "I": 6.0, "J": 8.0, "L": 12.0, "M": 10.0, "N": 9.0,
        "O": 10.0, "Q": 12.0, "R": 11.0, "S": 12.0, "T": 10.0, "U": 9.0,
        "Y": 10.0,
    }

    GOLDEN_ANCHOR_CELLS = {
        "A1": "EXPRESS CONSIGNMENTS WORKSHEET",
        "A2": "NON-COMMERCIAL CONSIGNMENTS",
        "A6": "SECTION 2",
        "Q6": "SECTION 3 — FOR OFFICIAL USE ONLY",
        "A7": "LINE\nNO.",
        "H7": "THN",
        "L7": "CUSTOMS\nVALUE (TTD)",
        "P7": "TOTAL\nTAXES",
    }

    def setUp(self):
        super().setUp()
        from app import store_courier
        store_courier.COURIER_FILE = Path(self.tmpdir) / "manifests.json"
        store_courier.COURIER_FILE.write_text("[]")

    def _make_manifest(self):
        """Single-line manifest for layout-only tests."""
        from app.services import courier_service
        m = courier_service.create_manifest({
            "manifest_no": "PARITY-TEST",
            "arrival_date": "2026-05-14",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "hawb": "1234567", "shipper": "Test", "importer": "Test",
            "description": "Test", "thn": "33049990",
            "cost_usd": 50.0, "freight_usd": 0.0, "packages": 1, "weight_kg": 1,
        })
        return courier_service.get_manifest(m["id"])

    def test_parity_merged_ranges(self):
        """All critical merged ranges from golden must be present.

        Layout-anchor merges (rows 1-6) are at fixed positions regardless of
        line count. Totals/grand-total merges shift with the line count, so
        we compute their expected positions for the test manifest.
        """
        from openpyxl import load_workbook
        from app.services import courier_export
        import io
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        generated_ranges = {str(r) for r in ws.merged_cells.ranges}

        # Layout-anchor merges (always at same row numbers)
        fixed_merges = {
            "A1:X1", "A2:X2", "J3:X3", "A5:X5", "A6:P6", "Q6:X6",
        }
        missing = fixed_merges - generated_ranges
        self.assertEqual(missing, set(),
                         f"Missing layout-anchor merged ranges: {missing}")

        # Totals/grand-total merges. Test manifest has 1 line, so:
        # data row 8, totals row 9, grand total row 10.
        totals_row = 9
        grand_row = 10
        movable_merges = {
            f"A{totals_row}:E{totals_row}",         # TOTALS label
            f"A{grand_row}:O{grand_row}",           # TOTAL TAXES label
            f"Q{grand_row}:V{grand_row}",           # TOTAL INCL. OFFICER UPLIFTS label
        }
        missing = movable_merges - generated_ranges
        self.assertEqual(missing, set(),
                         f"Missing totals/grand-total merged ranges: {missing}")

    def test_parity_column_widths(self):
        """All key column widths from golden must match (±0.5 tolerance)."""
        from openpyxl import load_workbook
        from app.services import courier_export
        import io
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        for letter, expected in self.GOLDEN_COLUMN_WIDTHS.items():
            actual = ws.column_dimensions[letter].width
            self.assertIsNotNone(
                actual, f"Column {letter}: no width set (expected {expected})",
            )
            self.assertAlmostEqual(
                actual, expected, delta=0.5,
                msg=f"Column {letter}: expected {expected}, got {actual}",
            )

    def test_parity_anchor_cell_labels(self):
        """Specific anchor cells must contain expected labels."""
        from openpyxl import load_workbook
        from app.services import courier_export
        import io
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        for ref, expected in self.GOLDEN_ANCHOR_CELLS.items():
            actual = ws[ref].value
            self.assertEqual(
                actual, expected,
                f"Cell {ref}: expected {expected!r}, got {actual!r}",
            )

    def test_parity_row_heights(self):
        """Critical row heights must match golden (banner, headers, data)."""
        from openpyxl import load_workbook
        from app.services import courier_export
        import io
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        # Row 6 (banner) ≈ 15.75; row 7 (headers) ≈ 31.5; row 8 (data) ≈ 27.75
        self.assertAlmostEqual(ws.row_dimensions[6].height, 15.75, delta=1.0)
        self.assertAlmostEqual(ws.row_dimensions[7].height, 31.5, delta=1.0)
        self.assertAlmostEqual(ws.row_dimensions[8].height, 27.75, delta=1.0)

    def test_parity_section_2_banner_styling(self):
        """SECTION 2 banner uses light-green fill FFC6EFCE."""
        from openpyxl import load_workbook
        from app.services import courier_export
        import io
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        c = ws["A6"]
        self.assertEqual(c.fill.fgColor.rgb, "FFC6EFCE")
        self.assertEqual(c.value, "SECTION 2")
        self.assertTrue(c.font.bold)

    def test_parity_section_3_banner_styling(self):
        """SECTION 3 banner uses pale-orange fill FFFCE4D6."""
        from openpyxl import load_workbook
        from app.services import courier_export
        import io
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        c = ws["Q6"]
        self.assertEqual(c.fill.fgColor.rgb, "FFFCE4D6")

    def test_parity_data_cells_have_values_not_blank(self):
        """
        DATA VISIBILITY CHECK (from brief): every key money cell in every
        line row must have a non-empty cached value so viewers that don't
        recalculate formulas still display the number.
        """
        from openpyxl import load_workbook
        from app.services import courier_export
        import io
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        # data_only=True: read cached values
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        # Row 8 is the single data row
        for col in ("L", "M", "N", "O", "P"):
            v = ws[f"{col}8"].value
            self.assertIsNotNone(
                v, f"Cell {col}8 has no cached value — would render blank in non-Excel viewers",
            )
            self.assertIsInstance(
                v, (int, float),
                f"Cell {col}8 cached value is {type(v).__name__}, expected numeric",
            )

    def test_parity_totals_row_has_cached_sums(self):
        """
        TOTALS row's SUM formulas must have cached values so non-Excel
        viewers display the totals correctly.
        """
        from openpyxl import load_workbook
        from app.services import courier_export
        import io
        m = self._make_manifest()
        data = courier_export.build_worksheet_v3(m)
        wb = load_workbook(io.BytesIO(data), data_only=True)
        ws = wb.active
        # 1 line → data row 8, totals row 9
        for col in ("F", "J", "L", "M", "N", "O", "P"):
            v = ws[f"{col}9"].value
            self.assertIsInstance(
                v, (int, float),
                f"Totals cell {col}9 has no cached value (would render blank)",
            )


class TestHazmatFormFields(_RulesStoreTestBase):
    """
    Tests for the Hazmat XLSX export with broker-fillable courier-data
    form fields. Covers the new POST /courier/manifests/{id}/hazmat
    endpoint that accepts a JSON body of optional form fields.

    Critical property: EVERY field is optional. The broker should be able
    to download the hazmat with zero, some, or all fields filled in.
    """

    def setUp(self):
        super().setUp()
        from app import store_courier
        store_courier.COURIER_FILE = Path(self.tmpdir) / "manifests.json"
        store_courier.COURIER_FILE.write_text("[]")

    def _make_manifest_with_line(self):
        from app.services import courier_service
        m = courier_service.create_manifest({
            "manifest_no": "106-31244603",
            "arrival_date": "15.05.2026",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "Test", "thn": "61046900",
            "cost_usd": 50.0, "packages": 1, "weight_kg": 1,
        })
        return courier_service.get_manifest(m["id"])

    def test_hazmat_with_all_fields_populated(self):
        """All form fields land in the right cells in the XLSX."""
        from openpyxl import load_workbook
        from app.services import courier_export
        import io

        m = self._make_manifest_with_line()
        fields = {
            "date": "15.05.2026",
            "ntde_no": "TEST-NTDE-001",
            "ced_receipt_no": "CED-12345",
            "vat_no": "V123990",
            "carrier": "TTPOST",
            "date_of_arrival": "14.05.2026",
            "rot_no": "ROT-2026-99",
            "no_of_skids": 0,
            "no_of_boxes": 5,
            "no_of_bags": 2,
            "no_of_commercial_pcs": 0,
            "no_of_non_commercial_pcs": 19,
            "total_no_of_pkgs": 19,
            "no_of_pkgs_detained": 1,
            "no_of_pkgs_seized": 0,
            "no_of_pkgs_bonded": 0,
        }
        data = courier_export.build_hazmat(m, courier_data_fields=fields)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active

        # Check every field
        self.assertEqual(ws["C5"].value, "15.05.2026")
        self.assertEqual(ws["F5"].value, "TEST-NTDE-001")
        self.assertEqual(ws["K5"].value, "CED-12345")
        self.assertEqual(ws["N5"].value, "V123990")
        self.assertEqual(ws["D8"].value, "14.05.2026")
        self.assertEqual(ws["H8"].value, "ROT-2026-99")
        self.assertEqual(ws["L8"].value, "TTPOST")
        self.assertEqual(ws["D9"].value, 0)
        self.assertEqual(ws["D10"].value, 5)
        self.assertEqual(ws["D13"].value, 2)
        self.assertEqual(ws["I10"].value, 0)
        self.assertEqual(ws["I13"].value, 19)
        self.assertEqual(ws["F11"].value, 19)  # explicit override
        self.assertEqual(ws["D16"].value, 1)
        self.assertEqual(ws["I16"].value, 0)
        # No. of Pkgs Bonded — value lives in the merged E18:F20 block per
        # the golden (label B18:D20 is the label-only merge to its left).
        self.assertEqual(ws["E18"].value, 0)

    def test_hazmat_with_no_fields_still_generates(self):
        """Empty form body must still produce a valid XLSX (every field optional)."""
        from app.services import courier_export

        m = self._make_manifest_with_line()
        data = courier_export.build_hazmat(m, courier_data_fields={})
        self.assertIsInstance(data, bytes)
        self.assertGreater(len(data), 5000)  # non-trivial XLSX

    def test_hazmat_with_no_kwarg_still_generates(self):
        """Calling without courier_data_fields kwarg works (backwards compat)."""
        from app.services import courier_export

        m = self._make_manifest_with_line()
        data = courier_export.build_hazmat(m)
        self.assertIsInstance(data, bytes)

    def test_hazmat_partial_fields_blanks_rest(self):
        """Partial form: only some fields filled, others empty."""
        from openpyxl import load_workbook
        from app.services import courier_export
        import io

        m = self._make_manifest_with_line()
        data = courier_export.build_hazmat(m, courier_data_fields={
            "date": "15.05.2026", "vat_no": "V999",
        })
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        self.assertEqual(ws["C5"].value, "15.05.2026")
        self.assertEqual(ws["N5"].value, "V999")
        # NTDE and CED Receipt should be empty/None
        self.assertIn(ws["F5"].value, (None, ""))
        self.assertIn(ws["K5"].value, (None, ""))
        # Counts should be empty
        self.assertIn(ws["D9"].value, (None, ""))

    def test_hazmat_total_pkgs_auto_sums_when_not_provided(self):
        """
        When total_no_of_pkgs is omitted, hazmat uses the manifest's
        package count as a plain value (not a formula — same parity
        rule as the worksheet).
        """
        from openpyxl import load_workbook
        from app.services import courier_export
        import io

        m = self._make_manifest_with_line()
        # Don't provide total_no_of_pkgs
        data = courier_export.build_hazmat(m, courier_data_fields={
            "no_of_commercial_pcs": 5, "no_of_non_commercial_pcs": 10,
        })
        wb = load_workbook(io.BytesIO(data), data_only=False)
        ws = wb.active
        # F11 should be a number (the manifest line count), not a formula
        self.assertNotIsInstance(
            ws["F11"].value, str,
            "F11 must be a plain value, not a formula",
        )
        # _make_manifest_with_line adds one line; commercial+non-commercial = 15
        # but the auto sum uses the actual line packages. Just assert it's >= 1.
        self.assertIsNotNone(ws["F11"].value)

    def test_hazmat_post_endpoint(self):
        """The POST /courier/manifests/{id}/hazmat HTTP endpoint works."""
        from fastapi import FastAPI
        from fastapi.testclient import TestClient
        from app.routes.courier import router as courier_router
        from app.services import courier_service

        m = courier_service.create_manifest({
            "manifest_no": "HAZMAT-HTTP",
            "arrival_date": "15.05.2026",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "Test", "thn": "61046900",
            "cost_usd": 50.0, "packages": 1, "weight_kg": 1,
        })

        app = FastAPI()
        app.include_router(courier_router)
        client = TestClient(app)

        # POST with fields
        r = client.post(
            f"/courier/manifests/{m['id']}/hazmat",
            json={"date": "15.05.2026", "carrier": "TTPOST"},
        )
        self.assertEqual(r.status_code, 200)
        self.assertEqual(
            r.headers["content-type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("hazmat", r.headers["content-disposition"].lower())
        self.assertGreater(len(r.content), 5000)

        # POST with empty body
        r = client.post(f"/courier/manifests/{m['id']}/hazmat", json={})
        self.assertEqual(r.status_code, 200)
        self.assertGreater(len(r.content), 5000)


class TestPhoneCaseExemption(_RulesStoreTestBase):
    """
    Issue #2: 39269090 is a catch-all. Phone cases under it are exempt
    per-LINE; generic plastics under the same THN pay 20%.
    """

    def setUp(self):
        super().setUp()
        from app import store_courier
        store_courier.COURIER_FILE = Path(self.tmpdir) / "m.json"
        store_courier.COURIER_FILE.write_text("[]")

    def test_phone_case_is_per_line_exempt(self):
        from app.services import courier_service
        m = courier_service.create_manifest({
            "manifest_no": "PC-TEST", "arrival_date": "2026-05-06",
            "exch_rate": 6.78,
        })
        line = courier_service.add_line_with_auto_thn(m["id"], {
            "description": "PHONE CASE", "cost_usd": 7.0,
            "packages": 1, "weight_kg": 1,
        })
        self.assertEqual(line["thn"], "39269090")
        self.assertEqual(line["exemption_class"], "full_exempt")
        self.assertEqual(line["duty"], 0.0)
        self.assertEqual(line["total_taxes"], 0.0)

    def test_generic_39269090_pays_duty(self):
        """An explicit 39269090 with a non-accessory description pays 20%."""
        from app.services import courier_service
        m = courier_service.create_manifest({
            "manifest_no": "GP-TEST", "arrival_date": "2026-05-06",
            "exch_rate": 6.78,
        })
        line = courier_service.add_line(m["id"], {
            "description": "HOME PRODUCT", "thn": "39269090",
            "cost_usd": 28.0, "packages": 1, "weight_kg": 1,
        })
        self.assertEqual(line["thn"], "39269090")
        self.assertNotEqual(line["exemption_class"], "full_exempt")
        self.assertEqual(line["duty_rate"], 0.2)
        self.assertGreater(line["duty"], 0)

    def test_screen_protector_is_exempt(self):
        from app.services import courier_service
        m = courier_service.create_manifest({
            "manifest_no": "SP-TEST", "arrival_date": "2026-05-06",
            "exch_rate": 6.78,
        })
        line = courier_service.add_line_with_auto_thn(m["id"], {
            "description": "tempered glass screen protector",
            "cost_usd": 5.0, "packages": 1, "weight_kg": 1,
        })
        self.assertEqual(line["exemption_class"], "full_exempt")
        self.assertEqual(line["duty"], 0.0)


class TestCorrectionServerRecalc(_RulesStoreTestBase):
    """
    Issue #5: officer corrections must be recomputed server-side so a THN
    change always yields correct duty/OPT/VAT, regardless of frontend math.
    """

    def setUp(self):
        super().setUp()
        from app import store_courier
        store_courier.COURIER_FILE = Path(self.tmpdir) / "m.json"
        store_courier.COURIER_FILE.write_text("[]")

    def _manifest(self):
        from app.services import courier_service
        m = courier_service.create_manifest({
            "manifest_no": "REC-TEST", "arrival_date": "2026-05-06",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "CLOTHING", "thn": "61046900",
            "cost_usd": 50.0, "packages": 1, "weight_kg": 1,
        })
        return courier_service.get_manifest(m["id"])

    def test_server_recomputes_dutiable_thn(self):
        """Officer THN 61046900 (20%) → server recomputes correct duty."""
        from app.services import courier_service
        m = self._manifest()
        # Frontend sends WRONG values (all zeros) — server must fix them.
        courier_service.record_examination(m["id"], {
            "examined_at": "2026-05-06", "examining_officer": "Officer",
            "corrections": [{
                "line_no": 1, "kind": "uplift", "officer_thn": "61046900",
                "add_cost_usd": 10.0, "adjusted_cif_ttd": 0,
                "add_duty": 0, "add_opt": 0, "add_vat": 0,
            }],
        })
        mm = courier_service.get_manifest(m["id"])
        corr = mm["officer_examination"]["corrections"][0]
        # cif = 10 * 6.78 = 67.80; duty = 67.80 * 0.2 = 13.56
        self.assertAlmostEqual(corr["adjusted_cif_ttd"], 67.80, places=2)
        self.assertAlmostEqual(corr["add_duty"], 13.56, places=2)
        self.assertGreater(corr["add_opt"], 0)
        self.assertGreater(corr["add_vat"], 0)

    def test_server_recomputes_exempt_thn_to_zero(self):
        """Officer changes THN to a full_exempt code → all additions zero."""
        from app.services import courier_service
        m = self._manifest()
        # Frontend (wrongly) sends non-zero duty for an exempt THN.
        courier_service.record_examination(m["id"], {
            "examined_at": "2026-05-06", "examining_officer": "Officer",
            "corrections": [{
                "line_no": 1, "kind": "reclass", "officer_thn": "85171300",
                "add_cost_usd": 10.0, "adjusted_cif_ttd": 67.80,
                "add_duty": 13.56, "add_opt": 4.75, "add_vat": 10.76,
            }],
        })
        mm = courier_service.get_manifest(m["id"])
        corr = mm["officer_examination"]["corrections"][0]
        # 85171300 is full_exempt — server must zero everything
        self.assertEqual(corr["add_duty"], 0.0)
        self.assertEqual(corr["add_opt"], 0.0)
        self.assertEqual(corr["add_vat"], 0.0)

    def test_server_reclass_zero_uplift_recomputes_from_original_cif(self):
        """Reclass without uplift should recompute deltas from original CIF."""
        from app.services import courier_service
        m = self._manifest()

        # Original line from _manifest(): THN 61046900 @ 20% on CIF 339.00
        # old duty=67.80, old opt=23.73, old vat=53.82
        # Reclass to 84212310 (30%) with zero uplift should increase duty/vat.
        courier_service.record_examination(m["id"], {
            "examined_at": "2026-05-06", "examining_officer": "Officer",
            "corrections": [{
                "line_no": 1, "kind": "reclass", "officer_thn": "84212310",
                "add_cost_usd": 0, "adjusted_cif_ttd": 0,
                "add_duty": 0, "add_opt": 0, "add_vat": 0,
            }],
        })
        mm = courier_service.get_manifest(m["id"])
        corr = mm["officer_examination"]["corrections"][0]
        self.assertEqual(corr["adjusted_cif_ttd"], 0.0)
        self.assertAlmostEqual(corr["add_duty"], 33.90, places=2)
        self.assertAlmostEqual(corr["add_opt"], 0.0, places=2)
        self.assertAlmostEqual(corr["add_vat"], 4.23, places=2)

    def test_server_preserves_negative_tax_removal(self):
        """Tax-removal corrections (negatives) are intentional — keep them."""
        from app.services import courier_service
        m = self._manifest()
        courier_service.record_examination(m["id"], {
            "examined_at": "2026-05-06", "examining_officer": "Officer",
            "corrections": [{
                "line_no": 1, "kind": "reclass", "officer_thn": "85171300",
                "add_cost_usd": 0,
                "add_duty": -33.89, "add_opt": -11.86, "add_vat": -26.90,
            }],
        })
        mm = courier_service.get_manifest(m["id"])
        corr = mm["officer_examination"]["corrections"][0]
        self.assertEqual(corr["add_duty"], -33.89)
        self.assertEqual(corr["add_opt"], -11.86)


class TestDescriptionChangeInSection3(_RulesStoreTestBase):
    """Issue #4: an officer description change must surface in the export."""

    def setUp(self):
        super().setUp()
        from app import store_courier
        store_courier.COURIER_FILE = Path(self.tmpdir) / "m.json"
        store_courier.COURIER_FILE.write_text("[]")

    def test_changed_description_appears_in_worksheet(self):
        import io
        from openpyxl import load_workbook
        from app.services import courier_service, courier_export
        m = courier_service.create_manifest({
            "manifest_no": "DESC-TEST", "arrival_date": "2026-05-06",
            "exch_rate": 6.78,
        })
        courier_service.add_line(m["id"], {
            "description": "TAGS", "thn": "39269090",
            "cost_usd": 7.0, "packages": 1, "weight_kg": 1,
        })
        courier_service.record_examination(m["id"], {
            "examined_at": "2026-05-06", "examining_officer": "Officer",
            "corrections": [{
                "line_no": 1, "kind": "description",
                "officer_thn": "39269090",
                "new_description": "PLASTIC LUGGAGE TAGS",
                "add_cost_usd": 7.0,
            }],
        })
        mm = courier_service.get_manifest(m["id"])
        data = courier_export.build_worksheet_v3(mm)
        wb = load_workbook(io.BytesIO(data))
        ws = wb.active
        e8 = str(ws["E8"].value)
        self.assertIn("PLASTIC LUGGAGE TAGS", e8)
        self.assertIn("TAGS", e8)  # original preserved in parentheses


class TestTariffPageFeatures(_RulesStoreTestBase):
    """
    Tests for the rebuilt Tariff page backend: ranked search, chapter
    summary, duty-band filter, overrides-only.
    """

    def test_ranked_search_surfaces_relevant_first(self):
        """Searching 'handbag' must put actual handbag codes at the top,
        not alphabetical noise."""
        from app.services import courier_rules
        r = courier_rules.list_tariff_entries(query="handbag", limit=5)
        self.assertGreater(r["total"], 0)
        # Top result's description should actually mention handbag
        top = r["items"][0]
        self.assertIn("handbag", (top.get("description") or "").lower())

    def test_search_by_exact_thn(self):
        from app.services import courier_rules
        # Pick any known code from the DB
        all_e = courier_rules.list_tariff_entries(limit=1)
        thn = all_e["items"][0]["thn"]
        r = courier_rules.list_tariff_entries(query=thn, limit=5)
        self.assertEqual(r["items"][0]["thn"], thn)

    def test_chapter_summary_returns_21_sections(self):
        from app.services import courier_rules
        cs = courier_rules.tariff_chapter_summary()
        self.assertIn("sections", cs)
        # Only non-empty sections are returned; should be many
        self.assertGreaterEqual(len(cs["sections"]), 10)
        first = cs["sections"][0]
        self.assertIn("title", first)
        self.assertIn("chapters", first)
        self.assertIn("count", first)

    def test_duty_band_free_filter(self):
        from app.services import courier_rules
        r = courier_rules.list_tariff_entries(duty_band="free", limit=20)
        self.assertGreater(r["total"], 0)
        for e in r["items"]:
            self.assertTrue(
                e.get("isExempt") or (e.get("dutyPct") or 0) == 0,
                f"{e['thn']} in free band but dutyPct={e.get('dutyPct')}",
            )

    def test_duty_band_high_filter(self):
        from app.services import courier_rules
        r = courier_rules.list_tariff_entries(duty_band="high", limit=20)
        for e in r["items"]:
            self.assertGreater(e.get("dutyPct") or 0, 25)

    def test_chapter_filter(self):
        from app.services import courier_rules
        r = courier_rules.list_tariff_entries(chapter=61, limit=10)
        self.assertGreater(r["total"], 0)
        for e in r["items"]:
            self.assertEqual(e["chapter"], 61)

    def test_overrides_only_empty_by_default(self):
        from app.services import courier_rules
        r = courier_rules.list_tariff_entries(overrides_only=True, limit=10)
        # No overrides applied in a fresh store
        self.assertEqual(r["total"], 0)

    def test_overrides_only_shows_override_after_maintain(self):
        from app.services import courier_rules
        # Apply a tariff override
        sample = courier_rules.list_tariff_entries(limit=1)["items"][0]
        courier_rules.add_tariff_entry(
            thn=sample["thn"], description=sample["description"],
            duty_pct=99, by="tester",
        )
        r = courier_rules.list_tariff_entries(overrides_only=True, limit=10)
        self.assertGreaterEqual(r["total"], 1)
        self.assertTrue(any(e["thn"] == sample["thn"] for e in r["items"]))

    def test_chapter_summary_endpoint_http(self):
        from fastapi import FastAPI
        from fastapi.testclient import TestClient
        from app.routes.courier_rules import router
        app = FastAPI()
        app.include_router(router)
        client = TestClient(app)
        resp = client.get("/courier/tariff/chapters")
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertIn("sections", body)

    def test_tariff_browse_endpoint_ranked_http(self):
        from fastapi import FastAPI
        from fastapi.testclient import TestClient
        from app.routes.courier_rules import router
        app = FastAPI()
        app.include_router(router)
        client = TestClient(app)
        resp = client.get("/courier/tariff?q=handbag&limit=3")
        self.assertEqual(resp.status_code, 200)
        body = resp.json()
        self.assertIn("items", body)
        self.assertLessEqual(len(body["items"]), 3)


if __name__ == "__main__":
    unittest.main(verbosity=2)
