"""
Courier Worksheet v3 + Hazmat XLSX generators.

Layouts are matched against the real broker templates:
  Worksheet_106-31245034_FINAL_v3.xlsx  (AWB 5034 reference)
  Courier_Data_Form_Hazmat_5034_v2.xlsx (AWB 5034 reference)

Worksheet layout
----------------
Title row 1:   "EXPRESS CONSIGNMENTS WORKSHEET"   (merged A1:X1)
Row 2:         "NON-COMMERCIAL CONSIGNMENTS"      (merged A2:X2)
Row 3:         A3 cargo reporter,  J3:X3 master waybill
Row 4:         A4 VAT no.,         F4 R.O.E.,    J4 freight
Row 5:         CBTT note (long)    (merged A5:X5)
Row 6:         A6:P6 "SECTION 2",  Q6:X6 "SECTION 3 — FOR OFFICIAL USE ONLY"
Row 7:         Column headers (multi-line)
Row 8+:        Data rows
TOTALS row:    after data, columns F G J L M N O P R S T U V W use SUM
Row +1:        "TOTAL TAXES ==>" merged A:O,  P=P{totals},
               "TOTAL INCL. OFFICER UPLIFTS ==>" merged Q:V,  W=P{totals}+W{totals}
Signature row: after a blank row

Section 2 columns: A B C D E F G H I J K L M N O P
                   line hawb shipper importer desc pkgs wt thn rate cost frgt cif duty opt vat total
Section 3 columns: Q R S T U V W X Y
                   officer-thn add-cost adj-cif add-duty add-opt add-vat add-total det/seized t-shed

Hazmat layout (Swissport Transit Shed)
--------------------------------------
This is NOT a per-line table. It is a transit shed manifest summary
with Trade and Non-Trade sections. Each section has three sub-rows:
Original Declared / Additional Taxes / Final Assessed. The Additional
row is a formula = Final - Original.

Tax columns at row 22:  F=CIF  H=OPT  J=DUTY  L=VAT  N=TOTAL
Trade section:          rows 23, 25 (formula), 27  (typically all zero for TTPOST)
Non-Trade section:      rows 31, 33 (formula), 35  (the manifest values)
Summary rows 38, 41, 42: copies of additional/final to dedicated summary lines.
"""
from __future__ import annotations

import io
import logging
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger("stallion.courier.export")


# ── Style tokens ─────────────────────────────────────────────────────────────
#
# These match the broker's golden template
# (Worksheet_106-31245034_FINAL_v3). The palette is light-green / pale-orange
# (TTPOST broker convention), not the navy/blue palette used internally before.
# Fonts are Calibri, sized to match the golden cell-for-cell.

FONT_TITLE = Font(name="Calibri", size=12, bold=True)
FONT_SUBTITLE = Font(name="Calibri", size=11, bold=True)
FONT_META = Font(name="Calibri", size=9, bold=False)
FONT_META_BOLD = Font(name="Calibri", size=9, bold=True)
FONT_NOTE = Font(name="Calibri", size=8, italic=True, color="FF595959")
FONT_BANNER = Font(name="Calibri", size=9, bold=True)
FONT_HEADER = Font(name="Calibri", size=8, bold=True)
FONT_BODY = Font(name="Calibri", size=9, color="FF000000")          # Section 2 data
FONT_BODY_BOLD = Font(name="Calibri", size=9, bold=True, color="FF000000")
FONT_BODY_S3 = Font(name="Calibri", size=9)                          # Section 3 data (no explicit color)
FONT_BODY_S3_BOLD = Font(name="Calibri", size=9, bold=True)
FONT_TOTALS_LABEL = Font(name="Calibri", size=9, bold=True)
FONT_SIGNATURE = Font(name="Calibri", size=9)

FILL_TITLE = PatternFill(fill_type=None)  # no fill on golden title rows
FILL_S2 = PatternFill("solid", start_color="FFC6EFCE")      # SECTION 2 light green
FILL_S3 = PatternFill("solid", start_color="FFFCE4D6")      # SECTION 3 pale orange
FILL_HEADER = PatternFill("solid", start_color="FFABEBC6")  # Section 2 column headers
FILL_HEADER_S3 = PatternFill("solid", start_color="FFFADBD8")  # Section 3 column headers (pink)
FILL_DATA = PatternFill("solid", start_color="FFFFFFFF")    # data row white (Section 2)
FILL_DATA_S3 = PatternFill("solid", start_color="FFFFF2CC")  # data row pale yellow (Section 3)
FILL_TOTALS = PatternFill("solid", start_color="FFABEBC6")  # totals row, same as header
FILL_HAZMAT_HEADER = PatternFill("solid", start_color="FFFFCC")   # pale yellow (golden)
FILL_HAZMAT_FINAL = PatternFill("solid", start_color="F2F2F2")    # light gray highlight (golden)

THIN = Side(style="thin", color="808080")
MEDIUM = Side(style="medium", color="808080")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_HEADER_S2 = Border(left=THIN, right=THIN, top=MEDIUM, bottom=MEDIUM)
BORDER_HEADER_S3 = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)   # for headers
ALIGN_CENTER_NOWRAP = Alignment(horizontal="center", vertical="center", wrap_text=False)
ALIGN_LEFT_WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)   # description data
ALIGN_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=False)
ALIGN_META = Alignment(horizontal="general", vertical="bottom", wrap_text=False)

FMT_TTD = "#,##0.00"
FMT_USD = "#,##0.00"
FMT_GENERAL = "General"


# ── Worksheet v3 layout ──────────────────────────────────────────────────────
# 25 columns (A–Y). Section 2 = A–P, Section 3 = Q–Y.
# Multi-line headers use \n; row height for the header row is 31.5pt
# to accommodate the wrapped text.

LAYOUT_V3_S2: List[Dict[str, Any]] = [
    {"col": "A", "label": "LINE\nNO.",        "width": 6.0,  "fmt": FMT_GENERAL, "align": ALIGN_CENTER_NOWRAP},
    {"col": "B", "label": "HAWB",             "width": 10.0, "fmt": FMT_GENERAL, "align": ALIGN_LEFT},
    {"col": "C", "label": "SHIPPER",          "width": 20.0, "fmt": FMT_GENERAL, "align": ALIGN_LEFT},
    {"col": "D", "label": "NAME OF\nIMPORTER","width": 22.0, "fmt": FMT_GENERAL, "align": ALIGN_LEFT},
    {"col": "E", "label": "DESCRIPTION OF GOODS", "width": 35.0, "fmt": FMT_GENERAL, "align": ALIGN_LEFT_WRAP},
    {"col": "F", "label": "NO.\nPKGS",        "width": 5.0,  "fmt": FMT_GENERAL, "align": ALIGN_CENTER_NOWRAP},
    {"col": "G", "label": "WT\n(lbs)",        "width": None, "fmt": FMT_GENERAL, "align": ALIGN_CENTER_NOWRAP},
    {"col": "H", "label": "THN",              "width": 12.0, "fmt": FMT_GENERAL, "align": ALIGN_CENTER_NOWRAP},
    {"col": "I", "label": "RATE",             "width": 6.0,  "fmt": FMT_GENERAL, "align": ALIGN_CENTER_NOWRAP},
    {"col": "J", "label": "COST\n(USD)",      "width": 8.0,  "fmt": FMT_USD,     "align": ALIGN_RIGHT},
    {"col": "K", "label": "FREIGHT",          "width": None, "fmt": FMT_USD,     "align": ALIGN_RIGHT},
    {"col": "L", "label": "CUSTOMS\nVALUE (TTD)", "width": 12.0, "fmt": FMT_TTD, "align": ALIGN_RIGHT},
    {"col": "M", "label": "DUTY",             "width": 10.0, "fmt": FMT_TTD,     "align": ALIGN_RIGHT},
    {"col": "N", "label": "OPT\n7%",          "width": 9.0,  "fmt": FMT_TTD,     "align": ALIGN_RIGHT},
    {"col": "O", "label": "VAT\n12.5%",       "width": 10.0, "fmt": FMT_TTD,     "align": ALIGN_RIGHT},
    {"col": "P", "label": "TOTAL\nTAXES",     "width": None, "fmt": FMT_TTD,     "align": ALIGN_RIGHT},
]

LAYOUT_V3_S3: List[Dict[str, Any]] = [
    {"col": "Q", "label": "OFFICER\nTHN",      "width": 12.0, "fmt": FMT_GENERAL, "align": ALIGN_CENTER_NOWRAP},
    {"col": "R", "label": "ADD.\nCOST (USD)",  "width": 11.0, "fmt": FMT_USD, "align": ALIGN_RIGHT},
    {"col": "S", "label": "ADJUSTED\nCIF (TTD)","width": 12.0, "fmt": FMT_TTD,"align": ALIGN_RIGHT},
    {"col": "T", "label": "ADD.\nDUTY",        "width": 10.0, "fmt": FMT_TTD, "align": ALIGN_RIGHT},
    {"col": "U", "label": "ADD.\nOPT",         "width": 9.0,  "fmt": FMT_TTD, "align": ALIGN_RIGHT},
    {"col": "V", "label": "ADD.\nVAT",         "width": None, "fmt": FMT_TTD, "align": ALIGN_RIGHT},
    {"col": "W", "label": "ADD.\nTOTAL",       "width": None, "fmt": FMT_TTD, "align": ALIGN_RIGHT},
    {"col": "X", "label": "DETAINED/\nSEIZED", "width": None, "fmt": FMT_GENERAL, "align": ALIGN_CENTER_NOWRAP},
    {"col": "Y", "label": "DEP. IN\nT/SHED",   "width": 10.0, "fmt": FMT_GENERAL, "align": ALIGN_CENTER_NOWRAP},
    {"col": "Z", "label": "NEW\nDESCRIPTION",  "width": 22.0, "fmt": FMT_GENERAL, "align": ALIGN_LEFT_WRAP},
]

# Row layout
ROW_TITLE_1 = 1     # "EXPRESS CONSIGNMENTS WORKSHEET"
ROW_TITLE_2 = 2     # "NON-COMMERCIAL CONSIGNMENTS"
ROW_META_1 = 3      # cargo reporter / master waybill
ROW_META_2 = 4      # VAT / R.O.E. / freight
ROW_NOTE = 5        # CBTT formula explainer
ROW_BANNER = 6      # "SECTION 2" / "SECTION 3 — FOR OFFICIAL USE ONLY"
ROW_HEADERS = 7     # column headers
ROW_DATA_START = 8  # first data row

LAST_MERGE_COL = "X"  # title bar / banners merge to column X (not Y)


# ── Worksheet v3 builders ────────────────────────────────────────────────────


def _set_widths(ws: Worksheet) -> None:
    """Apply column widths. None means leave at Excel's default (~8.43)."""
    for item in LAYOUT_V3_S2 + LAYOUT_V3_S3:
        if item.get("width") is not None:
            ws.column_dimensions[item["col"]].width = item["width"]


def _write_title_block(ws: Worksheet, manifest: Dict[str, Any]) -> None:
    """Rows 1–5: title, sub-title, meta block, formula explainer.

    All styling matches the broker's golden template
    (Worksheet_106-31245034_FINAL_v3): Calibri throughout, no fills on
    title rows, light-italic note row.
    """
    rate = float(manifest.get("exch_rate", 0))
    manifest_no = manifest.get("manifest_no", "")
    cargo_reporter = manifest.get("cargo_reporter", "TRINIDAD AND TOBAGO POSTAL CORPORATION")
    if cargo_reporter.upper() == "TTPOST":
        cargo_reporter = "TRINIDAD AND TOBAGO POSTAL CORPORATION"

    # Row 1 — main title
    ws.merge_cells(f"A{ROW_TITLE_1}:{LAST_MERGE_COL}{ROW_TITLE_1}")
    c = ws[f"A{ROW_TITLE_1}"]
    c.value = "EXPRESS CONSIGNMENTS WORKSHEET"
    c.font = FONT_TITLE
    c.alignment = ALIGN_CENTER_NOWRAP
    ws.row_dimensions[ROW_TITLE_1].height = 13.5

    # Row 2 — sub-title
    ws.merge_cells(f"A{ROW_TITLE_2}:{LAST_MERGE_COL}{ROW_TITLE_2}")
    c = ws[f"A{ROW_TITLE_2}"]
    c.value = "NON-COMMERCIAL CONSIGNMENTS"
    c.font = FONT_SUBTITLE
    c.alignment = ALIGN_CENTER_NOWRAP
    ws.row_dimensions[ROW_TITLE_2].height = 13.5

    # Row 3 — cargo reporter (left) and master waybill (right)
    ws[f"A{ROW_META_1}"] = f"CARGO REPORTER: {cargo_reporter}"
    ws[f"A{ROW_META_1}"].font = FONT_META
    ws[f"A{ROW_META_1}"].alignment = ALIGN_META

    ws.merge_cells(f"J{ROW_META_1}:{LAST_MERGE_COL}{ROW_META_1}")
    ws[f"J{ROW_META_1}"] = f"MASTER WAY BILL NUMBER: {manifest_no}"
    ws[f"J{ROW_META_1}"].font = FONT_META_BOLD
    ws[f"J{ROW_META_1}"].alignment = ALIGN_META
    ws.row_dimensions[ROW_META_1].height = 13.5

    # Row 4 — VAT no, R.O.E., freight
    vat_no = manifest.get("declarant_vat_no") or "V117369"
    ws[f"A{ROW_META_2}"] = f'VAT NO. / "N" NO.: {vat_no}'
    ws[f"F{ROW_META_2}"] = f"R.O.E.: {rate}"
    ws[f"J{ROW_META_2}"] = "FREIGHT:"
    for col, font in [("A", FONT_META), ("F", FONT_META_BOLD), ("J", FONT_META)]:
        ws[f"{col}{ROW_META_2}"].font = font
        ws[f"{col}{ROW_META_2}"].alignment = ALIGN_META
    ws.row_dimensions[ROW_META_2].height = 15.0

    # Row 5 — CBTT note (formula explainer)
    note = (
        f"CBTT RATE: USD 1.00 = TT$ {rate}  |  "
        "CIF = Cost x Rate  |  ICD = CIF x Duty%  |  "
        "OPT = CIF x 7%  |  VAT = (CIF+ICD+OPT) x 12.5%  |  "
        "NOTE: THN 8517.13.00 (smartphones) assessed FREE of all taxes under breakout exemption code"
    )
    ws.merge_cells(f"A{ROW_NOTE}:{LAST_MERGE_COL}{ROW_NOTE}")
    c = ws[f"A{ROW_NOTE}"]
    c.value = note
    c.font = FONT_NOTE
    c.alignment = ALIGN_META
    ws.row_dimensions[ROW_NOTE].height = 15.0


def _write_banner(ws: Worksheet) -> None:
    """Row 6 — SECTION 2 (light green) and SECTION 3 (pale orange) banners."""
    ws.merge_cells(f"A{ROW_BANNER}:P{ROW_BANNER}")
    c = ws[f"A{ROW_BANNER}"]
    c.value = "SECTION 2"
    c.font = FONT_BANNER
    c.fill = FILL_S2
    c.alignment = ALIGN_CENTER_NOWRAP

    ws.merge_cells(f"Q{ROW_BANNER}:{LAST_MERGE_COL}{ROW_BANNER}")
    c = ws[f"Q{ROW_BANNER}"]
    c.value = "SECTION 3 — FOR OFFICIAL USE ONLY"
    c.font = FONT_BANNER
    c.fill = FILL_S3
    c.alignment = ALIGN_CENTER_NOWRAP

    ws.row_dimensions[ROW_BANNER].height = 15.75


def _write_column_headers(ws: Worksheet) -> None:
    """
    Row 7 — multi-line column headers.

    Section 2 (A:P) uses light-green fill and medium top/bottom borders
    (golden's emphasis). Section 3 (Q:Y) uses pale-pink fill with thin
    borders all around. A7 specifically uses thin all around (no medium).
    """
    for item in LAYOUT_V3_S2 + LAYOUT_V3_S3:
        c = ws[f"{item['col']}{ROW_HEADERS}"]
        c.value = item["label"]
        c.font = FONT_HEADER
        c.alignment = ALIGN_CENTER
        # Per-column fill + border
        if item["col"] == "A":
            c.fill = FILL_HEADER
            c.border = BORDER_THIN
        elif item["col"] in "BCDEFGHIJKLMNOP":
            c.fill = FILL_HEADER
            c.border = BORDER_HEADER_S2
        else:  # Section 3: Q-Y
            c.fill = FILL_HEADER_S3
            c.border = BORDER_HEADER_S3
    ws.row_dimensions[ROW_HEADERS].height = 31.5


def _rate_display(line: Dict[str, Any]) -> str:
    """Map line classification to the rate cell text."""
    cls = line.get("exemption_class", "none")
    rate = float(line.get("duty_rate") or 0)
    if cls == "full_exempt":
        return "FREE"  # real worksheet uses FREE for fully-exempt cell phones
    if cls == "duty_free_only":
        return "FREE"
    if rate > 0:
        return f"{int(round(rate * 100))}%"
    return "FREE"


def _compute_line_values(line: Dict[str, Any], rate: float) -> Dict[str, float]:
    """
    Return the tax values for a line EXACTLY as stored on the line.

    The Stallion workbench computes and stores `cif_ttd`, `duty`, `opt`,
    `vat`, `total_taxes` on each line (rounded to 2 d.p. — what the broker
    sees on screen). We write those exact stored numbers into the XLSX so
    the worksheet matches the workbench cell-for-cell with zero drift.

    We do NOT recompute here. Recomputing in the exporter (even with the
    same formula) reintroduces floating-point divergence from what the
    broker saw. The line is the single source of truth.
    """
    cif = float(line.get("cif_ttd") or 0)
    duty = float(line.get("duty") or 0)
    opt = float(line.get("opt") or 0)
    vat = float(line.get("vat") or 0)
    total = line.get("total_taxes")
    if total is None:
        total = duty + opt + vat
    return {
        "cif": cif,
        "duty": duty,
        "opt": opt,
        "vat": vat,
        "total": float(total),
    }


def _compute_cached_totals(
    manifest: Dict[str, Any], lines: List[Dict[str, Any]], rate: float,
) -> Dict[str, float]:
    """
    Compute the cached numeric values for the TOTALS row's SUM formulas.

    Returns a mapping from column letter to its server-side computed sum.
    """
    out = {col: 0.0 for col in
           ("F", "G", "J", "L", "M", "N", "O", "P",
            "R", "S", "T", "U", "V", "W")}

    for line in lines:
        v = _compute_line_values(line, rate)
        out["F"] += int(line.get("packages") or 1)
        out["G"] += int(line.get("weight_kg") or 0)
        out["J"] += float(line.get("cost_usd") or 0)
        out["L"] += v["cif"]
        out["M"] += v["duty"]
        out["N"] += v["opt"]
        out["O"] += v["vat"]
        out["P"] += v["total"]

    # Officer uplifts / corrections / discovered lines.
    # NOTE: officer-discovered lines have line_no == None but STILL carry
    # additional taxes that must roll into the Section 3 totals.
    exam = manifest.get("officer_examination") or {}
    for corr in exam.get("corrections", []):
        out["R"] += float(corr.get("add_cost_usd") or 0)
        adj = corr.get("adjusted_cif_ttd")
        if adj is None:
            adj = float(corr.get("add_cost_usd") or 0) * rate
        out["S"] += float(adj)
        out["T"] += float(corr.get("add_duty") or 0)
        out["U"] += float(corr.get("add_opt") or 0)
        out["V"] += float(corr.get("add_vat") or 0)
        add_total = corr.get("add_total")
        if add_total is None:
            add_total = (float(corr.get("add_duty") or 0)
                         + float(corr.get("add_opt") or 0)
                         + float(corr.get("add_vat") or 0))
        out["W"] += float(add_total)

    return out


def _set_with_formula(cell, formula: str, cached_value: float, number_format: str) -> None:
    """
    Write the exact computed value to the cell. NO FORMULA.

    Formulas were removed entirely because Excel recalculates them on open,
    and the recalculated result drifts from the value the broker saw in the
    Stallion workbench (floating-point rounding, rate precision, order of
    operations all differ between our Python math and Excel's). The broker
    needs the worksheet to show EXACTLY the figures from the workbench, so
    we populate plain numeric values.

    The `formula` argument is intentionally ignored — kept only so the many
    call sites don't all need to change. The cell becomes a static number.
    """
    _ = formula  # deliberately unused — no formulas in generated sheets
    cell.value = cached_value
    cell.number_format = number_format


def _write_data_row(
    ws: Worksheet,
    row: int,
    line: Dict[str, Any],
    rate: float,
    correction: Optional[Dict[str, Any]] = None,
) -> None:
    """
    Write one Section 2 + Section 3 data row matching the golden template.

    Tax formulas mirror the broker's golden worksheet exactly:
      L = J*<rate>      (CIF)
      M = L*<duty_pct>  (Duty)
      N = L*0.07        (OPT)
      O = (L+M+N)*0.125 (VAT)
      P = M+N+O         (Total)
      W = T+U+V         (Add Total — Section 3)

    Each formula also gets a server-computed cached value so the cell
    displays correctly in non-Excel viewers (Apple Preview, Google Sheets
    web viewer, mobile Excel without recalc).
    """
    # Section 2 — identification
    ws[f"A{row}"] = int(line.get("line_no") or 0)
    # HAWB: keep numeric if it's all digits (golden stores it as int)
    raw_hawb = line.get("hawb", "")
    if isinstance(raw_hawb, (int, float)):
        ws[f"B{row}"] = int(raw_hawb)
    elif isinstance(raw_hawb, str) and raw_hawb.isdigit():
        ws[f"B{row}"] = int(raw_hawb)
    else:
        ws[f"B{row}"] = str(raw_hawb or "")
    ws[f"C{row}"] = line.get("shipper", "")
    ws[f"D{row}"] = line.get("importer", "")
    # Description: if the officer changed it during examination, show the
    # new description with the original in parentheses so the change is
    # visible on the worksheet (Issue #4). Otherwise just the declared desc.
    base_desc = line.get("description", "")
    if correction and (correction.get("new_description") or "").strip():
        new_desc = correction["new_description"].strip()
        if new_desc and new_desc != base_desc:
            ws[f"E{row}"] = f"{new_desc}  (was: {base_desc})"
        else:
            ws[f"E{row}"] = base_desc
    else:
        ws[f"E{row}"] = base_desc
    ws[f"F{row}"] = int(line.get("packages") or 1)
    ws[f"G{row}"] = int(line.get("weight_kg") or 0)  # column header says "lbs"; value as-stored
    ws[f"H{row}"] = str(line.get("thn", ""))
    ws[f"I{row}"] = _rate_display(line)
    ws[f"J{row}"] = float(line.get("cost_usd") or 0)
    ws[f"J{row}"].number_format = FMT_USD  # money format even before _set_with_formula loop
    if line.get("freight_usd"):
        ws[f"K{row}"] = float(line["freight_usd"])
        ws[f"K{row}"].number_format = FMT_USD

    # Compute cached values for the formula cells
    values = _compute_line_values(line, rate)

    cls = line.get("exemption_class", "none")
    duty_rate = float(line.get("duty_rate") or 0)

    # L = CIF. Golden uses =J{row}*<rate>; we use the same to match exactly.
    # If the broker types a value into K (freight), they'll need to update L
    # manually — same constraint as the broker's existing workflow.
    _set_with_formula(ws[f"L{row}"], f"=J{row}*{rate}", values["cif"], FMT_TTD)

    if cls == "full_exempt":
        ws[f"M{row}"] = 0
        ws[f"N{row}"] = 0
        ws[f"O{row}"] = 0
        ws[f"M{row}"].number_format = FMT_TTD
        ws[f"N{row}"].number_format = FMT_TTD
        ws[f"O{row}"].number_format = FMT_TTD
    elif cls == "duty_free_only":
        ws[f"M{row}"] = 0
        ws[f"M{row}"].number_format = FMT_TTD
        _set_with_formula(ws[f"N{row}"], f"=L{row}*0.07", values["opt"], FMT_TTD)
        _set_with_formula(ws[f"O{row}"], f"=(L{row}+M{row}+N{row})*0.125", values["vat"], FMT_TTD)
    else:
        _set_with_formula(ws[f"M{row}"], f"=L{row}*{round(duty_rate, 4)}", values["duty"], FMT_TTD)
        _set_with_formula(ws[f"N{row}"], f"=L{row}*0.07", values["opt"], FMT_TTD)
        _set_with_formula(ws[f"O{row}"], f"=(L{row}+M{row}+N{row})*0.125", values["vat"], FMT_TTD)

    _set_with_formula(ws[f"P{row}"], f"=M{row}+N{row}+O{row}", values["total"], FMT_TTD)

    # Section 3 — officer columns
    if correction:
        ws[f"Q{row}"] = str(correction.get("officer_thn", ""))
        ws[f"R{row}"] = float(correction.get("add_cost_usd") or 0)
        adj = correction.get("adjusted_cif_ttd")
        if adj is not None:
            ws[f"S{row}"] = float(adj)
        else:
            adj_value = float(correction.get("add_cost_usd") or 0) * rate
            _set_with_formula(ws[f"S{row}"], f"=R{row}*{rate}", adj_value, FMT_TTD)
        ws[f"T{row}"] = float(correction.get("add_duty") or 0)
        ws[f"U{row}"] = float(correction.get("add_opt") or 0)
        ws[f"V{row}"] = float(correction.get("add_vat") or 0)
        add_total = float(correction.get("add_duty") or 0) + \
                    float(correction.get("add_opt") or 0) + \
                    float(correction.get("add_vat") or 0)
        _set_with_formula(ws[f"W{row}"], f"=T{row}+U{row}+V{row}", add_total, FMT_TTD)
        if correction.get("detained_seized"):
            ws[f"X{row}"] = "Yes"
        # Y is "DEP. IN T/SHED" (transit shed timestamp, broker fills).
        # New description column is Z (added so Stallion preserves a
        # description change from the officer without displacing T/SHED).
        ws[f"Z{row}"] = str(correction.get("new_description") or "")

    # Apply per-cell formatting matching the golden:
    #   - Section 2 (A:P): alternating row fill (white on even data rows,
    #     FFF2F2F2 light gray on odd data rows starting from row 9).
    #   - Section 3 (Q:Z): solid pale-yellow fill (no alternating).
    #   - Section 2 data cells use FONT_BODY (color FF000000); P column is
    #     always bold ("Total Taxes" column emphasis in golden).
    #   - Section 3 data cells use FONT_BODY_S3 (no explicit color); W
    #     column is bold when a correction is present.
    #   - All money cells get #,##0.00 even when empty.
    SECTION_3_MONEY_COLS = {"R", "S", "T", "U", "V", "W"}
    SECTION_2_MONEY_COLS = {"J", "K", "L", "M", "N", "O", "P"}
    # Alternating fill: data starts at ROW_DATA_START (=8). Even data-row
    # offsets (relative to ROW_DATA_START) → white; odd → light gray.
    is_alt = ((row - ROW_DATA_START) % 2 == 1)
    s2_fill = PatternFill("solid", start_color="FFF2F2F2") if is_alt else FILL_DATA
    for item in LAYOUT_V3_S2 + LAYOUT_V3_S3:
        c = ws[f"{item['col']}{row}"]
        c.alignment = item["align"]
        c.border = BORDER_THIN
        # Section 2 (A:P) → alternating; Section 3 (Q:Y) → solid yellow
        if item["col"] in "QRSTUVWXY":
            c.fill = FILL_DATA_S3
            # W is bold when there's a correction; otherwise plain
            if item["col"] == "W" and correction:
                c.font = FONT_BODY_S3_BOLD
            else:
                c.font = FONT_BODY_S3
        else:
            c.fill = s2_fill
            # P (Total Taxes) is always bold per golden
            if item["col"] == "P":
                c.font = FONT_BODY_BOLD
            else:
                c.font = FONT_BODY
        # Apply number format. Money cells always get #,##0.00.
        if item["col"] in SECTION_3_MONEY_COLS or item["col"] in SECTION_2_MONEY_COLS:
            c.number_format = FMT_TTD
        else:
            c.number_format = item["fmt"]

    ws.row_dimensions[row].height = 27.75


def _write_totals_row(
    ws: Worksheet, row: int, first: int, last: int,
    cached_totals: Dict[str, float],
) -> None:
    """
    Write the TOTALS row matching the golden template.

    Golden styling:
      - A{row}:E{row} merged with "TOTALS" label, light-green fill,
        general (bottom) alignment.
      - F G J L M N O P (Section 2) hold SUM formulas with light-green
        fill (FFABEBC6).
      - R S T U V W (Section 3) hold SUM formulas with pale-yellow fill
        (FFFFF2CC).
      - All totals cells use bold Calibri without explicit font color,
        right-aligned, with #,##0.00 number format.
    """
    ws.merge_cells(f"A{row}:E{row}")
    ws[f"A{row}"] = "TOTALS"
    ws[f"A{row}"].font = FONT_TOTALS_LABEL
    ws[f"A{row}"].fill = FILL_TOTALS
    ws[f"A{row}"].alignment = ALIGN_META

    # SUM formula cells with cached values.
    # Golden uses #,##0.00 for ALL SUM cells (even the package/weight counts).
    for col in ("F", "G", "J", "L", "M", "N", "O", "P",
                "R", "S", "T", "U", "V", "W"):
        cached = cached_totals.get(col, 0.0)
        _set_with_formula(
            ws[f"{col}{row}"],
            f"=SUM({col}{first}:{col}{last})",
            cached,
            FMT_TTD,
        )

    # Apply golden-matching styling. Section 2 (A-P) → light green, no
    # borders. Section 3 R-W → pale yellow with thin borders. Q, X, Y →
    # pale yellow with no border, no alignment.
    NO_BORDER = Border()
    SECTION_3_SUM_COLS = {"R", "S", "T", "U", "V", "W"}
    for item in LAYOUT_V3_S2 + LAYOUT_V3_S3:
        c = ws[f"{item['col']}{row}"]
        if item["col"] == "A":
            c.font = FONT_TOTALS_LABEL
            c.fill = FILL_TOTALS
            c.border = NO_BORDER
        elif item["col"] in "BCDEFGHIJKLMNOP":
            c.font = FONT_BODY_S3_BOLD  # no explicit color
            c.fill = FILL_TOTALS
            c.border = NO_BORDER
            # Only F G J L M N O P get right alignment (data money cells)
            if item["col"] in "FGJLMNOP":
                c.alignment = ALIGN_RIGHT
        elif item["col"] in SECTION_3_SUM_COLS:
            c.font = FONT_BODY_S3_BOLD
            c.fill = FILL_DATA_S3
            c.border = BORDER_THIN
            c.alignment = ALIGN_RIGHT
        else:  # Q, X, Y
            c.font = FONT_BODY_S3_BOLD
            c.fill = FILL_DATA_S3
            c.border = NO_BORDER

    ws.row_dimensions[row].height = 18.0


def _write_grand_total_row(
    ws: Worksheet, row: int, totals_row: int,
    cached_total_taxes: float, cached_total_inc_uplifts: float,
) -> None:
    """
    Grand total row after the TOTALS row:
      A{row}:O{row} merged: 'TOTAL TAXES ==>' (no fill, bold, no color)
      P{row} = P{totals_row} (formula+cached, bold, no fill)
      Q{row}:V{row} merged: 'TOTAL INCL. OFFICER UPLIFTS ==>'
          (pale-pink fill FFFADBD8, bold, no color)
      W{row} = P{totals_row} + W{totals_row}
          (pale-pink fill + thin borders, bold, no color)
    """
    ws.merge_cells(f"A{row}:O{row}")
    ws[f"A{row}"] = "TOTAL TAXES ==>"
    ws[f"A{row}"].font = FONT_BODY_S3_BOLD  # bold without explicit color
    ws[f"A{row}"].alignment = ALIGN_META

    _set_with_formula(ws[f"P{row}"], f"=P{totals_row}", cached_total_taxes, FMT_TTD)
    ws[f"P{row}"].font = FONT_BODY_S3_BOLD  # no explicit color
    ws[f"P{row}"].alignment = ALIGN_RIGHT

    ws.merge_cells(f"Q{row}:V{row}")
    ws[f"Q{row}"] = "TOTAL INCL. OFFICER UPLIFTS ==>"
    ws[f"Q{row}"].font = FONT_BODY_S3_BOLD
    ws[f"Q{row}"].fill = FILL_HEADER_S3  # pale pink
    ws[f"Q{row}"].alignment = ALIGN_META

    _set_with_formula(
        ws[f"W{row}"], f"=P{totals_row}+W{totals_row}",
        cached_total_inc_uplifts, FMT_TTD,
    )
    ws[f"W{row}"].font = FONT_BODY_S3_BOLD
    ws[f"W{row}"].fill = FILL_HEADER_S3  # pale pink
    ws[f"W{row}"].border = BORDER_THIN
    ws[f"W{row}"].alignment = ALIGN_RIGHT

    ws.row_dimensions[row].height = 18.0


def _write_signature_row(ws: Worksheet, row: int) -> None:
    """Final signature row — signature placeholders for officer + declarant."""
    ws[f"A{row}"] = "Signature of Examining Officer: ______________________________"
    ws[f"A{row}"].font = FONT_SIGNATURE
    ws[f"A{row}"].alignment = ALIGN_META

    ws[f"J{row}"] = "SIGNATURE, NAME & LICENCE NO OF DECLARANT: ______________________________"
    ws[f"J{row}"].font = FONT_SIGNATURE
    ws[f"J{row}"].alignment = ALIGN_META

    ws.row_dimensions[row].height = 15.0


def _correction_for(manifest: Dict[str, Any], line_no: int) -> Optional[Dict[str, Any]]:
    exam = manifest.get("officer_examination") or {}
    for c in exam.get("corrections", []):
        if c.get("line_no") == line_no:
            return c
    return None


def build_worksheet_v3(manifest: Dict[str, Any]) -> bytes:
    """Build the Worksheet v3 XLSX as bytes, matching the real broker template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    rate = float(manifest.get("exch_rate", 0))
    if rate <= 0:
        raise ValueError("exch_rate must be > 0")

    # Page setup: portrait, fit-to-page, A4 — matches real broker template.
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.3
    ws.page_margins.right = 0.3
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    _set_widths(ws)
    _write_title_block(ws, manifest)
    _write_banner(ws)
    _write_column_headers(ws)

    lines = manifest.get("lines", []) or []
    for i, line in enumerate(lines):
        row = ROW_DATA_START + i
        correction = _correction_for(manifest, line.get("line_no"))
        _write_data_row(ws, row, line, rate, correction)

    if lines:
        first_data = ROW_DATA_START
        last_data = ROW_DATA_START + len(lines) - 1
        totals_row = last_data + 1
        grand_row = totals_row + 1
        sig_row = grand_row + 2

        # Compute cached SUM values server-side so the worksheet renders
        # correctly in viewers that don't recalculate formulas.
        cached_totals = _compute_cached_totals(manifest, lines, rate)

        _write_totals_row(ws, totals_row, first_data, last_data, cached_totals)
        _write_grand_total_row(
            ws, grand_row, totals_row,
            cached_total_taxes=cached_totals.get("P", 0.0),
            cached_total_inc_uplifts=cached_totals.get("P", 0.0) + cached_totals.get("W", 0.0),
        )
        _write_signature_row(ws, sig_row)

    # Officer-discovered new lines (line_no=null) appended below signature
    exam = manifest.get("officer_examination") or {}
    new_corrs = [c for c in exam.get("corrections", []) if c.get("line_no") is None]
    if new_corrs:
        anchor_row = (ROW_DATA_START + len(lines) + 4) if lines else ROW_DATA_START
        ws.merge_cells(f"A{anchor_row}:{LAST_MERGE_COL}{anchor_row}")
        c = ws[f"A{anchor_row}"]
        c.value = "OFFICER-DISCOVERED LINES (Section 3 only)"
        c.font = FONT_BANNER
        c.fill = FILL_S3
        c.alignment = ALIGN_CENTER
        for i, corr in enumerate(new_corrs):
            r = anchor_row + 1 + i
            ws[f"Q{r}"] = corr.get("officer_thn", "")
            ws[f"R{r}"] = float(corr.get("add_cost_usd") or 0)
            adj = corr.get("adjusted_cif_ttd")
            ws[f"S{r}"] = float(adj) if adj is not None else f"=R{r}*{rate}"
            ws[f"T{r}"] = float(corr.get("add_duty") or 0)
            ws[f"U{r}"] = float(corr.get("add_opt") or 0)
            ws[f"V{r}"] = float(corr.get("add_vat") or 0)
            ws[f"W{r}"] = f"=T{r}+U{r}+V{r}"
            for item in LAYOUT_V3_S3:
                cell = ws[f"{item['col']}{r}"]
                cell.font = FONT_BODY
                cell.alignment = item["align"]
                cell.border = BORDER_THIN
                if item["fmt"] not in ("@", None):
                    cell.number_format = item["fmt"]
            ws.row_dimensions[r].height = 27.75

    # Freeze the top header rows so scrolling keeps them visible
    ws.freeze_panes = f"B{ROW_DATA_START}"

    buf = io.BytesIO()
    wb.save(buf)

    # No formula injection needed any more — every computed cell holds a
    # plain numeric value (see _set_with_formula). What the broker sees in
    # the workbench is written verbatim into the XLSX with no recalculation.
    return buf.getvalue()


# ── Hazmat (Swissport Transit Shed Form) ─────────────────────────────────────
# This is a manifest-level form, NOT a per-line block layout. It contains
# Trade and Non-Trade sections with three sub-rows each:
#   - Original Declared on Worksheet
#   - Additional Taxes (formula = Final - Original)
#   - Final Assessed Values
#
# For TTPOST the Trade section is all-zero by convention (TTPOST express
# is non-commercial).
#
# Tax columns at row 22:  F=CIF  H=OPT  J=DUTY  L=VAT  N=TOTAL

HAZMAT_TAX_COLS = {
    "cif":   "F",
    "opt":   "H",
    "duty":  "J",
    "vat":   "L",
    "total": "N",
}

# Trade section (typically zero for TTPOST express)
HAZMAT_TRADE_ORIGINAL = 23
HAZMAT_TRADE_ADDITIONAL = 25
HAZMAT_TRADE_FINAL = 27

# Non-Trade section (the manifest values land here)
HAZMAT_NT_ORIGINAL = 31
HAZMAT_NT_ADDITIONAL = 33
HAZMAT_NT_FINAL = 35

# Summary footer
HAZMAT_TOTAL_ADDITIONAL = 38   # = Non-Trade Additional row (38 cells reference 33)
HAZMAT_TOTAL_TAXES = 41        # = Non-Trade Final row (41 cells reference 35)


def _hzm_sum_total(ws: Worksheet, row: int) -> None:
    """N column total for one tax row = F + H + J + L."""
    cols = HAZMAT_TAX_COLS
    ws[f"{cols['total']}{row}"] = (
        f"={cols['cif']}{row}+{cols['opt']}{row}+"
        f"{cols['duty']}{row}+{cols['vat']}{row}"
    )


def _hzm_apply_money_format(ws: Worksheet, rows: List[int]) -> None:
    for r in rows:
        for col in HAZMAT_TAX_COLS.values():
            cell = ws[f"{col}{r}"]
            cell.number_format = FMT_TTD
            cell.font = FONT_BODY
            cell.alignment = ALIGN_RIGHT
            cell.border = BORDER_THIN


def _hzm_arial(size: int = 11, bold: bool = False) -> Font:
    """Arial font matching the golden hazmat template."""
    return Font(name="Arial", size=size, bold=bold, color="FF000000")


def _hzm_calibri(size: int = 10, bold: bool = False) -> Font:
    """Value font for filled cells (golden uses Calibri 10 for data)."""
    return Font(name="Calibri", size=size, bold=bold)


def _hzm_format_date(value: Any, default: str = "") -> str:
    """
    Format a date as DD.MM.YYYY to match the golden form.
    Accepts ISO strings (YYYY-MM-DD) or DD.MM.YYYY (passed through).
    Falls back to whatever was supplied if parsing fails.
    """
    if value in (None, ""):
        return default
    s = str(value).strip()
    # Already in golden DD.MM.YYYY format
    if len(s) == 10 and s[2] == "." and s[5] == ".":
        return s
    # ISO YYYY-MM-DD
    if len(s) >= 10 and s[4] == "-" and s[7] == "-":
        return f"{s[8:10]}.{s[5:7]}.{s[:4]}"
    return s


def _write_hazmat_meta(
    ws: Worksheet, manifest: Dict[str, Any], fields: Optional[Dict[str, Any]] = None,
) -> None:
    """
    Top of the Hazmat form — exact reproduction of the broker's golden
    template (Courier_Data_Form_Hazmat_4614.xlsx).

    Layout: title block at B2:O4 (Arial 16 bold, no fill, medium border),
    then a meta block on rows 5-20 with merged label/value cells matching
    the golden exactly (see the row map in the rewrite commit message).

    `fields` (all optional, missing keys leave the cell blank):
      - date, ntde_no, ced_receipt_no, vat_no
      - carrier, date_of_arrival, rot_no
      - no_of_skids, no_of_boxes, no_of_bags
      - no_of_commercial_pcs, no_of_non_commercial_pcs, total_no_of_pkgs
      - no_of_pkgs_detained, no_of_pkgs_seized, no_of_pkgs_bonded
    """
    fields = fields or {}
    awb = manifest.get("manifest_no", "")
    arrival = manifest.get("arrival_date", "")
    cargo_reporter = manifest.get("cargo_reporter", "TTPOST")
    if cargo_reporter == "TRINIDAD AND TOBAGO POSTAL CORPORATION":
        cargo_reporter = "TTPOST"

    def field(name: str, default: Any = "") -> Any:
        v = fields.get(name)
        return v if v not in (None, "") else default

    # ── Title (B2:O4, Arial 16 bold, pale-yellow fill, medium border) ─────
    ws.merge_cells("B2:O4")
    title = ws["B2"]
    title.value = "SWISSPORT TRANSIT SHED COURIER DATA FORM"
    title.font = _hzm_arial(16, bold=True)
    title.alignment = ALIGN_CENTER
    title.fill = FILL_HAZMAT_HEADER  # pale yellow same as tax-header row
    # The medium border in the golden wraps the merged range; openpyxl
    # only applies the border to the top-left of a merge, but we paint
    # the border around the visible perimeter for visual parity.
    medium = Side(style="medium", color="FF000000")
    for c in "BCDEFGHIJKLMNO":
        ws[f"{c}2"].border = Border(top=medium)
        ws[f"{c}4"].border = Border(bottom=medium)
    ws["B2"].border = Border(top=medium, left=medium)
    ws["O2"].border = Border(top=medium, right=medium)
    ws["B3"].border = Border(left=medium)
    ws["O3"].border = Border(right=medium)
    ws["B4"].border = Border(bottom=medium, left=medium)
    ws["O4"].border = Border(bottom=medium, right=medium)
    ws.row_dimensions[2].height = 20.25
    ws.row_dimensions[3].height = 13.5
    ws.row_dimensions[4].height = 22.5

    label = _hzm_arial(11, bold=True)
    value = _hzm_calibri(10)
    big_value = _hzm_arial(14, bold=False)  # date cell C5 is 14pt in golden

    def put_label(coord: str, text: str, merge: str | None = None):
        cell = ws[coord]
        cell.value = text
        cell.font = label
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_THIN
        if merge:
            ws.merge_cells(merge)

    def put_value(coord: str, val: Any, merge: str | None = None,
                  font: Font = value):
        cell = ws[coord]
        cell.value = val
        cell.font = font
        cell.alignment = ALIGN_LEFT
        cell.border = BORDER_THIN
        if merge:
            ws.merge_cells(merge)

    # ── Row 5: Date / NTDE No / CED Receipt / VAT No ─────────────────────
    put_label("B5", "Date:", "B5:B6")
    # Date cell (C5:D6 area in golden — C5 only, but row 5-6 are tall)
    put_value("C5", _hzm_format_date(field("date"), _hzm_format_date(arrival)),
              font=big_value)
    put_label("E5", "NTDE No:", "E5:E6")
    put_value("F5", field("ntde_no"), "F5:H6", font=big_value)
    put_label("I5", "CED Receipt No.", "I5:J6")
    put_value("K5", field("ced_receipt_no"), "K5:L6", font=big_value)
    put_label("M5", "VAT No:", "M5:M6")
    put_value("N5", field("vat_no", manifest.get("declarant_vat_no") or "V117369"),
              "N5:O6", font=big_value)
    ws.row_dimensions[5].height = 14.25
    ws.row_dimensions[6].height = 14.25

    # ── Row 7: Name of courier / AWB ─────────────────────────────────────
    put_label("B7", "NAME OF COURIER:", "B7:C7")
    put_value("D7", cargo_reporter, "D7:I7")
    put_label("J7", "AWB/BL #")
    put_value("K7", awb, "K7:O7")
    ws.row_dimensions[7].height = 19.5

    # Row 8: Date of Arrival / Rot. No / Carrier.
    # Golden has D8 empty (the date is already shown in C5 above), so we
    # only write a value here if the broker explicitly supplied one. The
    # `date_of_arrival` field stays optional.
    put_label("B8", "Date of Arrival:", "B8:C8")
    doa = field("date_of_arrival")
    if doa:
        put_value("D8", _hzm_format_date(doa), "D8:F8")
    else:
        ws.merge_cells("D8:F8")
        ws["D8"].border = BORDER_THIN
    put_label("G8", "Rot. No")
    put_value("H8", field("rot_no"), "H8:J8")
    put_label("K8", "Carrier: ")  # trailing space matches golden
    put_value("L8", field("carrier"), "L8:O8")

    # ── Rows 9-20: package counts (laid out as in golden) ────────────────
    # Golden uses D-column for the value cell beside each label.
    put_label("B9", "No. of Skids.")
    sk = field("no_of_skids")
    if sk != "":
        put_value("D9", sk)
    else:
        ws["D9"].border = BORDER_THIN

    put_label("B10", "No. of Boxes.")
    put_value("D10", field("no_of_boxes"))
    put_label("G10", "No. of Commercial Pcs")
    cp = field("no_of_commercial_pcs")
    if cp != "":
        put_value("I10", cp)
    else:
        ws["I10"].border = BORDER_THIN

    # Row 11: Total No of Pkgs (label E11, value F11 — trailing space matches)
    put_label("E11", "Total No of Pkgs ")
    explicit_total = field("total_no_of_pkgs")
    n_pkgs_auto = sum(int(l.get("packages") or 0)
                      for l in manifest.get("lines", []))
    if explicit_total != "":
        put_value("F11", explicit_total)
    elif n_pkgs_auto:
        put_value("F11", n_pkgs_auto)
    else:
        ws["F11"].border = BORDER_THIN

    # Row 9-11: J9:K11 empty merged block (golden visual spacing)
    ws.merge_cells("J9:K11")
    ws["J9"].border = BORDER_THIN

    # Row 12: J12:K14 reserved (golden has empty merged block)
    ws.merge_cells("J12:K14")
    ws["J12"].value = " "
    ws["J12"].border = BORDER_THIN

    # Row 13: No. of bags / No. of Non Commercial Pcs
    put_label("B13", "No. of bags")
    put_value("D13", field("no_of_bags"))
    put_label("G13", "No. of Non Commercial Pcs")
    put_value("I13", field("no_of_non_commercial_pcs", n_pkgs_auto))

    # Other empty merged blocks visible on form (C11:D11, C14:D14,
    # E12:F14, E15:F17, J15:K17, J18:K20, G18:I20 are empty visual blocks
    for mr in ("C11:D11", "C14:D14", "E12:F14", "E15:F17",
               "J15:K17", "J18:K20", "G18:I20"):
        ws.merge_cells(mr)
        top_left = mr.split(":")[0]
        ws[top_left].border = BORDER_THIN

    # Row 16: No. of Pkgs Detained / Seized (golden has trailing space on
    # "Detained ")
    put_label("B16", "No. of Pkgs Detained ")
    put_value("D16", field("no_of_pkgs_detained"))
    put_label("G16", "No. of Pkgs Seized")
    put_value("I16", field("no_of_pkgs_seized"))

    # Row 18-20: No. of Pkgs Bonded — label in B18:D20, value in E18:F20.
    # The other right-side blocks (G18:I20, J18:K20) are empty form spaces.
    put_label("B18", "No. of Pkgs Bonded", "B18:D20")
    bonded = field("no_of_pkgs_bonded")
    ws.merge_cells("E18:F20")
    if bonded != "":
        ws["E18"].value = bonded
        ws["E18"].font = value
        ws["E18"].alignment = ALIGN_LEFT
    ws["E18"].border = BORDER_THIN


def _write_hazmat_tax_table(
    ws: Worksheet,
    manifest: Dict[str, Any],
) -> None:
    """
    Trade + Non-Trade tax sections — exact reproduction of the golden form.

    Every value cell is a 2-row × 2-col merged block (F23:G24, H23:I24,
    etc. in the golden). All money cells hold PLAIN VALUES, no formulas
    (same fix as the worksheet — Excel recalculation drifts from the
    workbench figures the broker saw). Highlight rows (Final Assessed,
    TOTAL TAXES, Total Additional Taxes) carry the pale-gray fill from
    the golden.
    """
    cols = HAZMAT_TAX_COLS
    totals = manifest.get("totals", {}) or {}

    label_font = _hzm_arial(12, bold=True)
    value_font = _hzm_calibri(10)

    # ── Row 22: tax column headers (pale yellow, Arial 12 bold, merged
    # as F22:G22, H22:I22, J22:K22, L22:M22, N22:O22) ────────────────────
    header_specs = [
        ("F22", "F22:G22", "CIF"),
        ("H22", "H22:I22", "OPT"),
        ("J22", "J22:K22", "DUTY"),
        ("L22", "L22:M22", "VAT"),
        ("N22", "N22:O22", "TOTAL"),
    ]
    for coord, merge, text in header_specs:
        ws.merge_cells(merge)
        cell = ws[coord]
        cell.value = text
        cell.font = label_font
        cell.fill = FILL_HAZMAT_HEADER
        cell.alignment = ALIGN_CENTER
        cell.border = BORDER_THIN
    ws.row_dimensions[22].height = 15.75

    def write_money_row(row: int, cif: float, opt: float, duty: float,
                        vat: float, total: float, highlight: bool = False,
                        blank_total: bool = False):
        """
        Write a money row across 5 merged 2-col blocks. Plain values, no
        formulas. `highlight` paints the gray fill seen on Final/TOTAL rows.
        `blank_total` matches the golden where Trade-zero rows leave the
        TOTAL column visually empty (the merge exists, the cell is blank).
        """
        merges = {
            "cif":   (f"F{row}",  f"F{row}:G{row+1}"),
            "opt":   (f"H{row}",  f"H{row}:I{row+1}"),
            "duty":  (f"J{row}",  f"J{row}:K{row+1}"),
            "vat":   (f"L{row}",  f"L{row}:M{row+1}"),
            "total": (f"N{row}",  f"N{row}:O{row+1}"),
        }
        values = {"cif": cif, "opt": opt, "duty": duty, "vat": vat,
                  "total": total}
        for key, (coord, merge) in merges.items():
            ws.merge_cells(merge)
            cell = ws[coord]
            if key == "total" and blank_total:
                cell.value = None
            else:
                cell.value = round(values[key], 2)
            cell.font = value_font
            cell.alignment = ALIGN_RIGHT
            cell.number_format = FMT_TTD
            cell.border = BORDER_THIN
            if highlight:
                cell.fill = FILL_HAZMAT_FINAL

    def write_section_label(row: int, line1: str, line2: str | None,
                            section_name: str | None, section_merge: str | None,
                            highlight_label_row: bool = False):
        """Write the B-column section labels and the E-column section name."""
        ws[f"B{row}"].value = line1
        ws[f"B{row}"].font = label_font
        ws[f"B{row}"].border = BORDER_THIN
        if highlight_label_row:
            for col in "BCDE":
                ws[f"{col}{row}"].fill = FILL_HAZMAT_FINAL
        if line2:
            ws[f"B{row+1}"].value = line2
            ws[f"B{row+1}"].font = label_font
            ws[f"B{row+1}"].border = BORDER_THIN
        if section_name and section_merge:
            ws.merge_cells(section_merge)
            ws[f"E{row}"].value = section_name
            ws[f"E{row}"].font = label_font
            ws[f"E{row}"].alignment = ALIGN_CENTER
            ws[f"E{row}"].border = BORDER_THIN

    # ── TRADE section (rows 23-28). Typically zero for TTPOST express. ──
    # Trade Original Declared (B23/B24 labels, E23:E28 section name "Trade")
    write_section_label(23, "Original Values Declared", "on Worksheet",
                        "Trade", "E23:E28")
    write_money_row(23, 0, 0, 0, 0, 0, blank_total=True)

    # Trade Additional Taxes (label on row 26, values on row 27).
    # Row 25 in the golden has empty merged blocks (F25:G26, H25:I26,
    # J25:K26, L25:M26, N25:O26) — purely visual spacing reserves. Create
    # them empty too for layout parity.
    for mr in ("F25:G26", "H25:I26", "J25:K26", "L25:M26", "N25:O26"):
        ws.merge_cells(mr)
        ws[mr.split(":")[0]].border = BORDER_THIN
    ws[f"B26"].value = "Additional Taxes"
    ws[f"B26"].font = label_font
    ws[f"B26"].border = BORDER_THIN
    write_money_row(27, 0, 0, 0, 0, 0, blank_total=True)

    # Trade Final Assessed (label on row 28, values would be on row 29 but
    # the golden actually has NO row 29 values — it has row 27 then jumps to
    # row 28 label. Looking at the golden again: row 23 values, row 26 label
    # "Additional Taxes", row 27 values, row 28 label "Final Assessed
    # Values" with NO values row beneath. So Trade has only 2 value rows.
    # The "Final Assessed Values" label highlights gray.
    ws[f"B28"].value = "Final Assessed Values"
    ws[f"B28"].font = label_font
    ws[f"B28"].border = BORDER_THIN
    for col in "BCD":
        ws[f"{col}28"].fill = FILL_HAZMAT_FINAL

    # ── NON-TRADE section (rows 31-36) — the manifest values land here ──
    orig_cif = float(totals.get("total_cif_ttd") or 0)
    orig_duty = float(totals.get("total_duty") or 0)
    orig_opt = float(totals.get("total_opt") or 0)
    orig_vat = float(totals.get("total_vat") or 0)
    orig_total = orig_cif + orig_opt + orig_duty + orig_vat
    # Wait — the golden's N31=1183.76 with F31=2819.61, H31=197.37,
    # J31=541.57, L31=444.82. Sum of OPT+DUTY+VAT = 197.37+541.57+444.82
    # = 1183.76. So the TOTAL column for the Non-Trade section is taxes
    # only (not including CIF). Match that.
    orig_total = orig_opt + orig_duty + orig_vat

    add_duty = 0.0
    add_opt = 0.0
    add_vat = 0.0
    add_cif = 0.0
    exam = manifest.get("officer_examination") or {}
    for c in exam.get("corrections", []) or []:
        add_duty += float(c.get("add_duty") or 0)
        add_opt += float(c.get("add_opt") or 0)
        add_vat += float(c.get("add_vat") or 0)
        add_cif += float(c.get("adjusted_cif_ttd") or 0)
    add_total = add_opt + add_duty + add_vat

    final_cif = orig_cif + add_cif
    final_opt = orig_opt + add_opt
    final_duty = orig_duty + add_duty
    final_vat = orig_vat + add_vat
    final_total = final_opt + final_duty + final_vat

    # Non-Trade Original (B31/B32 labels, E31:E36 section name "Non-Trade")
    write_section_label(31, "Original Values Declared", "on Worksheet",
                        "Non-Trade", "E31:E36")
    write_money_row(31, orig_cif, orig_opt, orig_duty, orig_vat, orig_total)

    # Non-Trade Additional (label on row 34, values on row 33)
    ws[f"B34"].value = "Additional Taxes"
    ws[f"B34"].font = label_font
    ws[f"B34"].border = BORDER_THIN
    write_money_row(33, add_cif, add_opt, add_duty, add_vat, add_total)

    # Non-Trade Final (label on row 36 with gray highlight, values on
    # row 35)
    ws[f"B36"].value = "Final Assessed Values"
    ws[f"B36"].font = label_font
    ws[f"B36"].border = BORDER_THIN
    for col in "BCD":
        ws[f"{col}36"].fill = FILL_HAZMAT_FINAL
    write_money_row(35, final_cif, final_opt, final_duty, final_vat, final_total)

    # ── Summary footer ──────────────────────────────────────────────────
    # Total Additional Taxes — label row 39, values row 38 (gray highlight
    # NOT shown on row 38 in golden — only row 41 highlights)
    ws[f"B39"].value = "Total Additional Taxes"
    ws[f"B39"].font = label_font
    ws[f"B39"].border = BORDER_THIN
    write_money_row(38, add_cif, add_opt, add_duty, add_vat, add_total)

    # TOTAL TAXES — label row 42 (gray), values row 41 (gray)
    ws[f"B42"].value = "TOTAL TAXES"
    ws[f"B42"].font = label_font
    ws[f"B42"].border = BORDER_THIN
    for col in "BCDE":
        ws[f"{col}42"].fill = FILL_HAZMAT_FINAL
    write_money_row(41, final_cif, final_opt, final_duty, final_vat,
                    final_total, highlight=True)

    # Row heights matching golden
    for r, h in [(23, 15.0), (24, 15.0), (25, 15.0), (26, 15.75),
                 (27, 15.0), (28, 15.0), (29, 15.0), (30, 15.0),
                 (31, 15.0), (32, 15.0), (33, 15.0), (34, 15.75),
                 (35, 15.0), (36, 15.0), (37, 15.0), (38, 15.0),
                 (39, 15.75), (41, 15.0), (42, 15.75)]:
        ws.row_dimensions[r].height = h


def _set_hazmat_widths(ws: Worksheet) -> None:
    # Widths from the golden template (Courier_Data_Form_Hazmat_4614.xlsx).
    # Columns G-O default — the golden only customises A-F and P-Q which
    # are off the visible form, so the rest fall to Excel default (~8.43).
    widths = {
        "A": 4.57, "B": 9.14, "C": 11.43, "D": 5.43, "E": 14.71, "F": 12.0,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


def build_hazmat(
    manifest: Dict[str, Any],
    courier_data_fields: Optional[Dict[str, Any]] = None,
) -> bytes:
    """
    Build the Swissport Transit Shed Courier Data Form XLSX as bytes.

    The output is an exact reproduction of the broker's golden template
    (Courier_Data_Form_Hazmat_4614.xlsx): same merges, fonts (Arial 11
    labels, Calibri 10 values, Arial 16 title), pale-yellow tax header
    row, gray-highlighted "Final Assessed Values" and "TOTAL TAXES"
    rows, landscape orientation. Every money cell holds a plain VALUE,
    no formulas — what the broker saw in the workbench is what the
    XLSX shows.

    `courier_data_fields` (all optional, missing fields render blank):
      - date, ntde_no, ced_receipt_no, vat_no
      - carrier, date_of_arrival, rot_no
      - no_of_skids, no_of_boxes, no_of_bags
      - no_of_commercial_pcs, no_of_non_commercial_pcs, total_no_of_pkgs
      - no_of_pkgs_detained, no_of_pkgs_seized, no_of_pkgs_bonded
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Landscape — matches golden.
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_options.horizontalCentered = True
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.5
    ws.page_margins.bottom = 0.5

    _set_hazmat_widths(ws)
    _write_hazmat_meta(ws, manifest, courier_data_fields or {})
    _write_hazmat_tax_table(ws, manifest)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Public surface ───────────────────────────────────────────────────────────

__all__ = [
    "build_worksheet_v3",
    "build_hazmat",
    "LAYOUT_V3_S2",
    "LAYOUT_V3_S3",
]
