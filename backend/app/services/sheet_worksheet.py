"""
sheet_worksheet.py — render a Stallion Sheet to the ACE-matched worksheet layout.

Matches the reference ACE/ASYCUDA worksheet:
  - EX-WORKS / INLAND / FOB / FREIGHT / INSURANCE / OTHER / CIF TT$ ladder
  - single CIF FACTOR (CIF TT$ / value), shown to 12 dp
  - each line CIF TT$ = line value * factor
  - DUTIES/TAXES summary split into RELIEF (R) and PAYABLE (P)

build(sheet, fmt="xlsx") -> (bytes, filename, media_type)
Drop into backend/app/services/sheet_worksheet.py
"""
from __future__ import annotations

import io
from typing import Any, Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

INK, GOLD, GREY, LINEG, AMBER = "1A1A1A", "B8860B", "E8E8E8", "F4F4F4", "FFF4D6"
_thin = Side(style="thin", color="BFBFBF")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _f(v, d=0.0):
    try:
        return float(v) if v not in (None, "") else d
    except (TypeError, ValueError):
        return d


def _set(c, val, *, bold=False, size=9, color=INK, fill=None, align="left",
         fmt=None, border=True, italic=False, wrap=False):
    c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = _border


def build(sheet: Dict[str, Any], fmt: str = "xlsx") -> Tuple[bytes, str, str]:
    exch = _f(sheet.get("exchange_rate"), 1.0)
    freight = _f(sheet.get("freight_usd"))
    insurance = _f(sheet.get("insurance_usd"))
    other = _f(sheet.get("other_usd"))
    inland = _f(sheet.get("inland_usd"))
    uplift = _f(sheet.get("uplift_pct"))
    cfu = _f(sheet.get("customs_user_fee"), 80.0)
    lines = sheet.get("lines", [])
    t = sheet.get("totals", {})
    relieved_mode = sheet.get("entry_mode") == "relieved"

    total_value = sum(_f(l.get("exworks_usd")) for l in lines) or 1.0
    fob_usd = total_value + inland + (total_value * uplift / 100)
    cif_usd = fob_usd + freight + insurance + other
    cif_ttd = round(cif_usd * exch, 2)
    factor = cif_ttd / total_value

    wb = Workbook()
    ws = wb.active
    ws.title = "Worksheet"

    ws.merge_cells("A1:L1")
    _set(ws["A1"], "W O R K S H E E T", bold=True, size=15, color=GOLD, border=False, align="center")
    ws.merge_cells("A2:L2")
    mode = "RELIEVED — Returning Resident" if relieved_mode else "Home Use — Duties & Taxes Payable"
    _set(ws["A2"], mode, italic=True, size=9, color="555555", border=False, align="center")
    ws.row_dimensions[1].height = 22

    # ── header grid (left ID block + right value ladder) ──
    left = [
        ("Consignee", sheet.get("consignee", "")),
        ("Consignor", sheet.get("consignor", "")),
        ("Vessel", sheet.get("vessel", "")),
        ("Bill of Lading", sheet.get("bl_number", "")),
        ("Rotation No.", sheet.get("rotation_no", "")),
        ("Invoice No.", sheet.get("invoice_no", "")),
        ("Invoice Date", sheet.get("invoice_date", "")),
        ("Port", sheet.get("port", "")),
        ("Date of Arrival", sheet.get("arrival_date", "")),
        ("Work Sheet Ref", sheet.get("reference", "")),
        ("Currency", sheet.get("currency", "USD")),
    ]
    ladder = [
        ("EX-WORKS (USD)", total_value),
        ("INLAND (USD)", inland),
        ("% UPLIFT", uplift),
        ("FOB (USD)", round(fob_usd, 2)),
        ("FREIGHT (USD)", freight),
        ("INSURANCE (USD)", insurance),
        ("OTHER (USD)", other),
        ("EXCHANGE RATE", exch),
        ("FREIGHT (TT$)", round(freight * exch, 2)),
        ("CIF TT$", cif_ttd),
        ("FACTOR", factor),
    ]
    r0 = 4
    for i in range(max(len(left), len(ladder))):
        r = r0 + i
        if i < len(left):
            _set(ws.cell(r, 1), left[i][0], bold=True, size=9, fill=GREY)
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
            _set(ws.cell(r, 2), left[i][1], size=9)
        if i < len(ladder):
            lbl, val = ladder[i]
            _set(ws.cell(r, 7), lbl, bold=True, size=9, fill=GREY)
            ws.merge_cells(start_row=r, start_column=8, end_row=r, end_column=12)
            is_factor = lbl == "FACTOR"
            _set(ws.cell(r, 8), val, size=9,
                 fmt='0.000000000000' if is_factor else ('0.0000' if lbl == "EXCHANGE RATE" else '#,##0.00'),
                 align="right", bold=is_factor,
                 fill=AMBER if is_factor else None, color="8A6D00" if is_factor else INK)
    r = r0 + max(len(left), len(ladder)) + 1

    # ── line table ──
    cols = ["#", "CPC", "ADD", "HS Code", "Description", "VALUE\nUSD", "CIF\nTT$",
            "DUTY %", "DUTY\nTT$", "VAT\nTT$", "R/P", "TOTAL TAX\nTT$"]
    widths = [4, 7, 6, 16, 30, 11, 13, 8, 11, 11, 6, 13]
    for i, (c, w) in enumerate(zip(cols, widths), start=1):
        _set(ws.cell(r, i), c, bold=True, size=8, color="FFFFFF", fill=INK, align="center", wrap=True)
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[r].height = 28

    first = r + 1
    for idx, ln in enumerate(lines, start=1):
        r += 1
        rf = LINEG if idx % 2 == 0 else None
        rel = bool(ln.get("relieved"))
        _set(ws.cell(r, 1), idx, size=8, align="center", fill=rf)
        _set(ws.cell(r, 2), ln.get("cpc", "4000"), size=8, align="center", fill=rf)
        _set(ws.cell(r, 3), "000", size=8, align="center", fill=rf)
        _set(ws.cell(r, 4), ln.get("hs_code", ""), size=8, align="center", fill=rf)
        _set(ws.cell(r, 5), ln.get("description", ""), size=8, fill=rf, wrap=True)
        _set(ws.cell(r, 6), _f(ln.get("exworks_usd")), size=8, align="right", fmt='#,##0.00', fill=rf, color="0000FF")
        # CIF TT$ = value * factor
        _set(ws.cell(r, 7), f"=ROUND(F{r}*$H${r0+10},2)", size=8, align="right", fmt='#,##0.00', fill=rf)
        _set(ws.cell(r, 8), _f(ln.get("duty_pct")), size=8, align="center", fmt='0.0', fill=rf)
        if rel:
            _set(ws.cell(r, 9), 0, size=8, align="right", fmt='#,##0.00', fill=rf)
            _set(ws.cell(r, 10), 0, size=8, align="right", fmt='#,##0.00', fill=rf)
            _set(ws.cell(r, 11), "R", size=8, align="center", fill=rf, color="1A5E3A", bold=True)
            _set(ws.cell(r, 12), 0, size=8, align="right", fmt='#,##0.00', fill=rf)
        else:
            _set(ws.cell(r, 9), f"=ROUND(G{r}*H{r}/100,2)", size=8, align="right", fmt='#,##0.00', fill=rf)
            _set(ws.cell(r, 10), f"=ROUND((G{r}+I{r})*{_f(ln.get('vat_pct'),12.5)}/100,2)", size=8, align="right", fmt='#,##0.00', fill=rf)
            _set(ws.cell(r, 11), "P", size=8, align="center", fill=rf, color="963A10", bold=True)
            _set(ws.cell(r, 12), f"=I{r}+J{r}", size=8, align="right", fmt='#,##0.00', fill=rf)
        ws.row_dimensions[r].height = 20
    last = r

    r += 1
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
    _set(ws.cell(r, 1), "WORKSHEET TOTALS", bold=True, size=9, fill=GREY, align="right")
    for col, letter in [(6, "F"), (7, "G"), (9, "I"), (10, "J"), (12, "L")]:
        _set(ws.cell(r, col), f"=SUM({letter}{first}:{letter}{last})", bold=True, size=9, align="right", fmt='#,##0.00', fill=GREY)
    _set(ws.cell(r, 8), "", fill=GREY)
    _set(ws.cell(r, 11), "", fill=GREY)
    totals_row = r

    # ── DUTIES/TAXES summary: RELIEF vs PAYABLE ──
    r += 2
    _set(ws.cell(r, 1), "DUTIES / TAXES SUMMARY", bold=True, size=10, color=GOLD, border=False)
    r += 1
    _set(ws.cell(r, 1), "DESCRIPTION", bold=True, size=8, color="FFFFFF", fill=INK)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws.cell(r, 7), "RELIEF (R)", bold=True, size=8, color="FFFFFF", fill=INK, align="right")
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    _set(ws.cell(r, 10), "PAYABLE (P)", bold=True, size=8, color="FFFFFF", fill=INK, align="right")
    ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)

    rd = _f(t.get("relief_duty")); rv = _f(t.get("relief_vat"))
    pd = _f(t.get("duty")); psur = _f(t.get("surcharge")); pv = _f(t.get("vat"))
    rows = [
        ("Import Duty", rd, pd),
        ("Surcharge", 0.0, psur),
        ("VAT", rv, pv),
        ("CFU  Customs User Fee", 0.0, cfu),
    ]
    for lbl, rval, pval in rows:
        r += 1
        _set(ws.cell(r, 1), lbl, size=9, fill=LINEG)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
        _set(ws.cell(r, 7), rval, size=9, align="right", fmt='#,##0.00')
        ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
        _set(ws.cell(r, 10), pval, size=9, align="right", fmt='#,##0.00')
        ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)
    r += 1
    _set(ws.cell(r, 1), "SUMMARY TOTALS", bold=True, size=10, color="FFFFFF", fill=INK)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    _set(ws.cell(r, 7), round(rd + rv, 2), bold=True, size=10, color="FFFFFF", fill=INK, align="right", fmt='#,##0.00')
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=9)
    _set(ws.cell(r, 10), round(pd + psur + pv + cfu, 2), bold=True, size=10, color="FFFFFF", fill=INK, align="right", fmt='#,##0.00')
    ws.merge_cells(start_row=r, start_column=10, end_row=r, end_column=12)

    buf = io.BytesIO()
    wb.save(buf)
    ref = (sheet.get("reference") or sheet.get("id") or "worksheet").replace("/", "-")
    return (buf.getvalue(), f"Stallion_Worksheet_{ref}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
