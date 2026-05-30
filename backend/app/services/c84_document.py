"""
c84_document.py — render a C84 duty/tax concession claim worksheet.

A C84 supports a customs entry by documenting the concession claimed: the
beneficiary, the qualifying basis, and the relief vs payable split per line.
This produces the broker-facing worksheet the client signs and that backs the
SAD; it is NOT the official ASYCUDA form, which is filed in ASYCUDA itself.

build(sheet) -> (bytes, filename, media_type)
Mirrors sheet_worksheet styling so the two documents read as one product.
"""
from __future__ import annotations

import io
from typing import Any, Dict, Tuple

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from . import concession_service

INK, GOLD, GREY, LINEG, AMBER = "1A1A1A", "B8860B", "E8E8E8", "F4F4F4", "FFF4D6"
GREEN, RED = "1A5C3A", "B02020"
_thin = Side(style="thin", color="BFBFBF")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_TTD = '#,##0.00;(#,##0.00);"-"'


def _f(v, d=0.0):
    try:
        return float(v) if v not in (None, "") else d
    except (TypeError, ValueError):
        return d


def _set(c, val, *, bold=False, size=9, color=INK, fill=None, align="left",
         fmt=None, border=True, italic=False, wrap=False):
    c.value = val
    c.font = Font(name="Arial", bold=bold, size=size, color=color, italic=italic)
    c.alignment = Alignment(horizontal=align, vertical="center", wrap_text=wrap)
    if fill:
        c.fill = PatternFill("solid", fgColor=fill)
    if fmt:
        c.number_format = fmt
    if border:
        c.border = _border


def build(sheet: Dict[str, Any]) -> Tuple[bytes, str, str]:
    conc = sheet.get("concession", {}) or {}
    lines = sheet.get("lines", [])
    primary_code = conc.get("code") or next(
        (l.get("concession_code") for l in lines if l.get("concession_code")), "")
    cat = concession_service.CONCESSION_BY_CODE.get(primary_code, {})

    wb = Workbook()
    ws = wb.active
    ws.title = "C84 Concession"
    # Label column A is narrow; B is the key label; C+ hold values / the line table.
    widths = [3, 15, 30, 13, 13, 13, 12, 13, 13]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    r = 1
    _set(ws.cell(r, 1), "C84 — DUTY / TAX CONCESSION CLAIM", bold=True, size=14,
         color=INK, border=False)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 1
    _set(ws.cell(r, 1), "Stallion · Trade Module — broker worksheet supporting the customs entry",
         italic=True, size=9, color="666666", border=False)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
    r += 2

    # ── Beneficiary / qualification block ──
    # Two key/value pairs per row: label in col 2 (or 6), value spans the rest.
    def kv(label, value, row, col):
        _set(ws.cell(row, col), label, bold=True, size=9, fill=LINEG, align="left")
        vcol = col + 1
        _set(ws.cell(row, vcol), value or "—", size=9)
        end = 5 if col == 2 else 9
        if vcol < end:
            ws.merge_cells(start_row=row, start_column=vcol, end_row=row, end_column=end)

    def section(title, row):
        _set(ws.cell(row, 1), title, bold=True, size=10, fill=GOLD, color="FFFFFF")
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)

    section("CONCESSION", r)
    r += 1
    kv("Concession", cat.get("label", primary_code or "—"), r, 2)
    kv("Quantum", (cat.get("quantum", "—")).upper(), r, 6)
    r += 1
    kv("Reference", sheet.get("reference", ""), r, 2)
    kv("C84 No.", conc.get("declaration_no", ""), r, 6)
    r += 1
    _set(ws.cell(r, 2), "Legal basis", bold=True, size=9, fill=LINEG)
    _set(ws.cell(r, 3), cat.get("legal", "") or "—", size=9)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=9)
    r += 2

    section("BENEFICIARY", r)
    r += 1
    kv("Name", conc.get("beneficiary_name", "") or sheet.get("consignee", ""), r, 2)
    kv("ID / Passport", conc.get("beneficiary_id", ""), r, 6)
    r += 1
    kv("Approval ref", conc.get("approval_ref", ""), r, 2)
    kv("Return date", conc.get("return_date", ""), r, 6)
    r += 1
    kv("Abroad from", conc.get("residence_abroad_from", ""), r, 2)
    kv("Abroad to", conc.get("residence_abroad_to", ""), r, 6)
    r += 2

    # ── Per-line relief table ──
    headers = ["#", "Description", "HS Code", "CIF TT$", "Full Duty",
               "Full VAT", "MVT", "Relieved", "Payable"]
    for c, h in enumerate(headers, start=1):
        _set(ws.cell(r, c), h, bold=True, size=9, fill=INK, color="FFFFFF",
             align="center" if c != 2 else "left")
    r += 1
    first_data = r
    for i, ln in enumerate(lines, start=1):
        _set(ws.cell(r, 1), i, align="center", size=9)
        _set(ws.cell(r, 2), ln.get("description", ""), size=9, wrap=True)
        _set(ws.cell(r, 3), ln.get("hs_code", ""), size=9, align="center")
        _set(ws.cell(r, 4), _f(ln.get("cif_ttd")), fmt=_TTD, align="right", size=9)
        _set(ws.cell(r, 5), _f(ln.get("full_duty")), fmt=_TTD, align="right", size=9)
        _set(ws.cell(r, 6), _f(ln.get("full_vat")), fmt=_TTD, align="right", size=9)
        _set(ws.cell(r, 7), _f(ln.get("mvt")) + _f(ln.get("relief_mvt")),
             fmt=_TTD, align="right", size=9)
        _set(ws.cell(r, 8), _f(ln.get("relief_total")), fmt=_TTD, align="right",
             size=9, color=GREEN, bold=_f(ln.get("relief_total")) > 0)
        _set(ws.cell(r, 9), _f(ln.get("total_tax")), fmt=_TTD, align="right",
             size=9, color=RED if _f(ln.get("total_tax")) > 0 else INK)
        r += 1
    last_data = r - 1

    # ── Totals row (formulas so the sheet recalculates) ──
    _set(ws.cell(r, 1), "", border=True)
    _set(ws.cell(r, 2), "TOTALS", bold=True, size=9, fill=GREY)
    _set(ws.cell(r, 3), "", fill=GREY)
    for col in (4, 5, 6, 7, 8, 9):
        rng = f"{chr(64 + col)}{first_data}:{chr(64 + col)}{last_data}"
        _set(ws.cell(r, col), f"=SUM({rng})", fmt=_TTD, align="right",
             bold=True, size=9, fill=GREY,
             color=GREEN if col == 8 else (RED if col == 9 else INK))
    r += 2

    # ── Summary box ──
    cfu = _f(sheet.get("customs_user_fee"), 80.0)
    _set(ws.cell(r, 7), "Customs User Fee", bold=True, size=9, fill=LINEG)
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    _set(ws.cell(r, 9), cfu, fmt=_TTD, align="right", size=9)
    r += 1
    _set(ws.cell(r, 7), "TOTAL PAYABLE", bold=True, size=10, fill=GOLD, color="FFFFFF")
    ws.merge_cells(start_row=r, start_column=7, end_row=r, end_column=8)
    _set(ws.cell(r, 9), f"=I{last_data + 1}+I{r - 1}", fmt=_TTD, align="right",
         bold=True, size=10, color=RED)
    r += 2

    _set(ws.cell(r, 1),
         "Relief figures are indicative. Vehicle caps must be confirmed against the "
         "current Customs notice before filing. This worksheet supports — and does "
         "not replace — the ASYCUDA C84.",
         italic=True, size=8, color="888888", border=False, wrap=True)
    ws.merge_cells(start_row=r, start_column=1, end_row=r + 1, end_column=9)
    ws.row_dimensions[r].height = 26

    buf = io.BytesIO()
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True if ws.sheet_properties.pageSetUpPr else None
    try:
        from openpyxl.worksheet.properties import PageSetupProperties
        ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    except Exception:
        pass
    wb.save(buf)
    ref = (sheet.get("reference") or sheet.get("id", "c84")).replace("/", "-")
    return (buf.getvalue(), f"C84_{ref}.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
