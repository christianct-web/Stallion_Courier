"""
Parse a TTPOST express-consignment Excel template and turn it into a manifest.

The TTPOST template ("Express Consignments Worksheet") looks like:

    Row 1: header "EXPRESS CONSIGNMENTS WORKSHEET"
    Row 2: subtitle "NON-COMMERCIAL CONSIGNMENTS"
    Row 3: A=cargo reporter, J=master waybill number
    Row 5: A=VAT no, I=freight
    Row 7: banner "SECTION 2"
    Row 8: banner "DETAILS OF ALL HOUSE WAYBILLS ISSUED..."
    Row 9: column headers (LINE NO. AWB | HAWB | SHIPPER | NAME OF IMPORTER | DESCRIPTION OF GOODS | NO. OF PKGS | WEIGHT | THN | RATE | COST | FREIGHT | CUSTOMS VALUE | DUTY | OPT | VAT | TOTAL TAXES)
    Row 10+: data rows. Each row has:
        A: line no (1, 2, 3, ...)
        B: HAWB
        C: shipper
        D: importer
        E: description
        F: packages
        G: weight
        H: THN (USUALLY EMPTY — the broker uses Stallion to classify)
        J: cost (USD)
        K: freight (often empty)

The parser:
    1. Extracts the manifest header (master waybill, cargo reporter, VAT no.)
    2. Iterates data rows from the first row after the column-headers row
    3. Stops when description+cost are both empty
    4. Returns a structured dict ready for courier_service.create_manifest + add_line_with_auto_thn

This module does NOT create manifests — it only parses. The route handler
calls courier_service to actually create the manifest with auto-classified THNs.
"""
from __future__ import annotations

import io
import re
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


# ── Helpers ──────────────────────────────────────────────────────────────────


def _to_str(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _to_int(v: Any, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


def _to_float(v: Any, default: float = 0.0) -> float:
    if v is None or v == "":
        return default
    try:
        return float(v)
    except (TypeError, ValueError):
        return default


# ── Header extraction ────────────────────────────────────────────────────────


# Match "MASTER WAY BILL NUMBER: 106-31244382" or similar variants
MASTER_WAYBILL_RE = re.compile(
    r"(?:MASTER\s*WAY[\s]*BILL[\s]*(?:NO|NUMBER)?[\s:.\-]*)?(106-\d{6,})",
    re.IGNORECASE,
)
VAT_NO_RE = re.compile(r"VAT[\s.]*NO[\s.]*[/]?[\s\"]*N[\"]?\s*NO[\s.:\s]*([\w\d]+)", re.IGNORECASE)


def _find_master_waybill(ws) -> str:
    """Look across the top rows for the master waybill number."""
    for row in range(1, 10):
        for col in range(1, 20):
            v = ws.cell(row=row, column=col).value
            if not isinstance(v, str):
                continue
            m = MASTER_WAYBILL_RE.search(v)
            if m:
                return m.group(1)
    return ""


def _find_cargo_reporter(ws) -> str:
    """Find the cargo reporter (usually 'TRINIDAD AND TOBAGO POSTAL CORPORATION')."""
    for row in range(1, 10):
        for col in range(1, 20):
            v = ws.cell(row=row, column=col).value
            if not isinstance(v, str):
                continue
            if "CARGO REPORTER" in v.upper():
                # Pull whatever follows the colon
                if ":" in v:
                    return v.split(":", 1)[1].strip()
                return v.strip()
    return "TTPOST"


def _find_vat_no(ws) -> str:
    """Find the broker's VAT / 'N' number."""
    for row in range(1, 10):
        for col in range(1, 20):
            v = ws.cell(row=row, column=col).value
            if not isinstance(v, str):
                continue
            m = VAT_NO_RE.search(v)
            if m:
                return m.group(1).strip()
    return ""


# ── Data row extraction ──────────────────────────────────────────────────────


# Headers we look for in the column-header row, to confirm we found the right row
EXPECTED_HEADERS = ("LINE NO", "HAWB", "SHIPPER", "IMPORTER", "DESCRIPTION")


def _find_header_row(ws) -> int:
    """Find the row index containing the column headers."""
    for row in range(1, min(ws.max_row + 1, 30)):
        # Concatenate all cell values in this row
        row_text = " | ".join(
            _to_str(ws.cell(row=row, column=col).value).upper()
            for col in range(1, 17)
        )
        # Need at least 3 of the expected headers
        matches = sum(1 for h in EXPECTED_HEADERS if h in row_text)
        if matches >= 3:
            return row
    return 9  # default to row 9 (TTPOST default)


def _iter_data_rows(ws, start_row: int):
    """
    Yield (row_index, cells) tuples for each data row.
    Stops when 3 consecutive rows have no description AND no cost.
    """
    blank_streak = 0
    for row in range(start_row, ws.max_row + 1):
        description = _to_str(ws.cell(row=row, column=5).value)  # E
        cost = ws.cell(row=row, column=10).value  # J
        cost_is_empty = cost is None or cost == ""

        if not description and cost_is_empty:
            blank_streak += 1
            if blank_streak >= 3:
                return
            continue

        blank_streak = 0
        yield row, {
            "line_no_raw": ws.cell(row=row, column=1).value,    # A
            "hawb": _to_str(ws.cell(row=row, column=2).value),  # B
            "shipper": _to_str(ws.cell(row=row, column=3).value),  # C
            "importer": _to_str(ws.cell(row=row, column=4).value), # D
            "description": description,                          # E
            "packages": _to_int(ws.cell(row=row, column=6).value, default=1),  # F
            "weight_kg": _to_float(ws.cell(row=row, column=7).value),  # G
            "thn": _to_str(ws.cell(row=row, column=8).value).replace(".", ""),  # H
            "cost_usd": _to_float(cost),                          # J
            "freight_usd": _to_float(ws.cell(row=row, column=11).value),  # K
        }


# ── Public API ───────────────────────────────────────────────────────────────


def parse_ttpost_template(file_bytes: bytes) -> Dict[str, Any]:
    """
    Parse a TTPOST express-consignment Excel template into a structured dict.

    Returns:
        {
            "manifest_no": "106-31244382",
            "cargo_reporter": "TRINIDAD AND TOBAGO POSTAL CORPORATION",
            "vat_no": "V117369",
            "lines": [
                {"hawb": "...", "shipper": "...", "importer": "...",
                 "description": "...", "packages": 1, "weight_kg": 1.0,
                 "thn": "" (or pre-filled), "cost_usd": 21.0,
                 "freight_usd": 0.0, "source_row": 10},
                ...
            ],
            "warnings": ["..."]  # parsing issues that didn't block import
        }

    Raises:
        ValueError: if the file is unreadable or has no recognizable structure.
    """
    try:
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
    except Exception as e:
        raise ValueError(f"Could not open Excel file: {e}")

    if not wb.sheetnames:
        raise ValueError("Excel file has no sheets")

    ws = wb.active

    warnings: List[str] = []

    manifest_no = _find_master_waybill(ws)
    if not manifest_no:
        warnings.append(
            "Could not find master waybill number in header (looked for pattern '106-XXXXXX'). "
            "You'll need to set it manually."
        )

    cargo_reporter = _find_cargo_reporter(ws)
    vat_no = _find_vat_no(ws)

    header_row = _find_header_row(ws)
    data_start = header_row + 1

    lines: List[Dict[str, Any]] = []
    for row_idx, row_data in _iter_data_rows(ws, data_start):
        # Skip lines that are clearly invalid (no description AND no cost)
        if not row_data["description"] and row_data["cost_usd"] == 0:
            continue
        # We don't need line_no_raw — courier_service will renumber.
        row_data.pop("line_no_raw", None)
        row_data["source_row"] = row_idx
        lines.append(row_data)

    if not lines:
        raise ValueError(
            f"No data rows found starting at row {data_start}. "
            f"Make sure the file is a TTPOST express-consignment worksheet."
        )

    return {
        "manifest_no": manifest_no,
        "cargo_reporter": cargo_reporter,
        "vat_no": vat_no,
        "header_row": header_row,
        "data_start_row": data_start,
        "lines": lines,
        "warnings": warnings,
    }


def summarize_parsed(parsed: Dict[str, Any]) -> Dict[str, Any]:
    """Compact summary for logging / UI feedback."""
    return {
        "manifest_no": parsed["manifest_no"],
        "lines": len(parsed["lines"]),
        "warnings": parsed["warnings"],
    }
